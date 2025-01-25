"""Common routines for creating CDR web forms.

This module has the following sections:

    Imports of modules on which this one depends

    Tools for creating CGI forms and reports

        FieldStorage
            Replacement for deprecated cgi.FieldStorage class
        Controller
            Base class for top-level controller of CGI scripts
        FormFieldFactory
            Provides convenient class methods for creating HTML form fields
        HTMLPage
            Creates a CDR web page for forms and reports using a common
            customizable layout; uses the USWDS framework
        Reporter
            Creates CDR reports which can be sent to the web browser or
            to Excel
        BasicWebPage
            Framework for creating reports which need more space than can
            be provided by the USWDS-based classes
        Excel
            Wrapper for building Excel workbooks using OpenPyXl
        AdvancedSearch
            Search forms for each of the CDR document types
"""

# Packages from the standard library.
from collections import OrderedDict, UserDict
from datetime import date, datetime
from email.utils import parseaddr as parse_email_address
from functools import cached_property
from json import load as load_json_file, loads as load_json_string
from os import environ
from pathlib import Path
from re import compile as re_compile, search as re_search, sub as re_sub
from string import hexdigits
from sys import argv as sys_argv, exit as sys_exit
from sys import stdin as sys_stdin, stdout as sys_stdout
from types import SimpleNamespace
from urllib import parse as urllib_parse

# Third-party libraries/packages.
from lxml import html as lxml_html
from lxml.html import builder as html_builder, HtmlElement
from multipart import MultipartError, MultipartParser, parse_options_header
from openpyxl import styles as excel_styles, Workbook as ExcelWorkbook

# Project modules.
from cdr import Board, getControlGroup, getDoctype, Logging, TMP, URDATE
from cdrapi import db
from cdrapi.docs import Doc
from cdrapi.settings import Tier
from cdrapi.users import Session


class FieldStorage:
    """Replacement for deprecated cgi.FieldStorage class."""

    ENCODED = "application/x-www-form-urlencoded", "application/x-url-encoded"

    def __init__(self, logger=None):
        """Parse the field values we get from the browser via the web server.

        Optional keyword argument:
            logger - pass this in to perform debug logging for the values
        """

        # Get some values from the environment.
        self.logger = logger
        if logger:
            logger.debug("Constructing FieldStorage object")
        method = environ.get("REQUEST_METHOD", "GET").upper()
        content_type = environ.get("CONTENT_TYPE", self.ENCODED[0])
        content_type, options = parse_options_header(content_type)
        charset = options.get("charset", "utf-8")
        self.__fields = {}
        query_string = environ.get("QUERY_STRING")
        qs_opts = dict(encoding=charset, keep_blank_values=True)

        # Capture values passed as parameters appended to the URL.
        if query_string:
            if logger:
                logger.debug("query_string: %s", query_string)
            for key, value in urllib_parse.parse_qsl(query_string, **qs_opts):
                if logger:
                    logger.debug("adding %s from query_string", key)
                self.__add(self.SimpleValue(key, value))

        # Read values POSTed by the client.
        if method not in ("GET", "HEAD"):
            content_length = int(environ.get("CONTENT_LENGTH", "-1"))
            if logger:
                logger.debug("content_length: %s", content_length)
                logger.debug("content_type: %s", content_type)
            if content_type == "multipart/form-data":
                boundary = options.get("boundary")
                if logger:
                    logger.debug("boundary: %s", boundary)
                if not boundary:
                    message = "No boundary for multipart/form-data"
                    raise MultipartError(message)
                args = sys_stdin.buffer, boundary, content_length
                kwargs = dict(charset=charset)
                for part in MultipartParser(*args, **kwargs):
                    if logger:
                        logger.debug("adding %s from multipart", part.name)
                    self.__add(self.StreamedValue(part))
            elif content_type in self.ENCODED and content_length > 0:
                data = sys_stdin.buffer.read(content_length).decode(charset)
                if logger:
                    logger.debug("parsing %s", data)
                for key, value in urllib_parse.parse_qsl(data, **qs_opts):
                    self.__add(self.SimpleValue(key, value))

    def __add(self, item):
        if item.name not in self.__fields:
            self.__fields[item.name] = []
        self.__fields[item.name].append(item)

    def getfirst(self, key, default=None):
        """Return the first value received."""

        items = self.__fields.get(key, [])
        return items[0].value if items else default

    def getlist(self, key):
        """Return list of received values."""
        return [item.value for item in self.__fields.get(key, [])]

    def getvalue(self, key, default=None):
        """Return single value, list of values, or None."""

        items = self.__fields.get(key, [])
        if not items:
            return default
        if len(items) == 1:
            return items[0].value
        return [item.value for item in items]

    def keys(self):
        """Dictionary-style keys() method."""
        return self.__fields.keys()

    def __bool__(self):
        """True if we found any fields else false."""
        return bool(self.__fields)

    def __contains__(self, key):
        """Dictionary-style __contains__() method."""
        return key in self.__fields

    def __getitem__(self, key):
        """Dictionary-style indexing, returning the value objects."""

        if key not in self.__fields:
            raise KeyError(key)
        items = self.__fields[key]
        return items[0] if len(items) == 1 else items

    def __iter__(self):
        return iter(self.keys())

    def __len__(self):
        """Dictionary-style len(x) support."""
        return len(self.keys())

    def __str__(self):
        """Improve debugging/logging output."""

        fields = []
        for name, values in self.__fields.items():
            if len(values) < 1:
                fields.append(f"{name}=''")
            elif len(values) == 1:
                fields.append(f"{name}={values[0]!r}")
            else:
                fields.append(f"{name}={values!r}")
        fields = ", ".join(fields)
        return f"[{fields}]"

    def __repr__(self) -> str:
        """Pass through the serialized string."""
        return str(self)

    class SimpleValue:
        """Basic name+value pairings."""

        def __init__(self, name, value):
            """Store the name and value and create stubs for the rest."""

            self.name = name
            self.value = value
            self.filename = self.file = None

        def __str__(self):
            """Show the value for debugging/logging."""
            return self.value

        def __repr__(self) -> str:
            """Improve debugging/logging output."""
            return repr(self.value)

    class StreamedValue():
        """Items which might be a posted file or other streamed value."""

        def __init__(self, part):
            """Store the object extracted by the Multipart parser."""
            self.__part = part

        def __str__(self):
            """Show the value for debugging/logging."""
            return f"file {self.filename}"

        def __repr__(self):
            """Show the value for debugging/logging."""
            return f"file {self.filename}"

        @cached_property
        def name(self):
            """The string for the name of the value."""
            return self.__part.name

        @cached_property
        def value(self):
            """String (or bytes if from a file) for the value."""
            return self.__part.raw if self.filename else self.__part.value

        @cached_property
        def filename(self):
            """String for the name (not the whole path) for a posted file."""
            return self.__part.filename

        @cached_property
        def file(self):
            """File handle from which the binary file content can be read."""
            return self.__part.file


class Controller:
    """Base class for top-level controller for a CGI script.

    Includes methods for displaying a form (typically for a report)
    and for rendering the requested report.

    This will gradually replace the older `Control` class, which
    is built around the use of `Page` objects, which created
    HTML pages using direct string manipulation instead of real
    HTML parser objects.

    Typically, a web script will create a derived class using this
    class as its base. It is possible to override the constructor,
    but this is not necessary. If you do, be sure to invoke the
    base class constructor from the overriding implementation.

    The `run()` method is the top-level processing starting point.
    That method calls the `show_report()` method if the Submit
    button was clicked, or `show_form()` if no button was clicked.
    The derived class implements `populate_form()` to add the
    fields it needs for its job, and overrides the `build_tables()`
    method to provide the report's data.
    """

    DATETIMELEN = len("YYYY-MM-DD HH:MM:SS")
    TAMPERING = "CGI parameter tampering detected"
    USERNAME = "UserName"
    PASSWORD = "Password"
    PORT = "Port"
    SESSION = "Session"
    REQUEST = "Request"
    DOCID = "DocId"
    BASE = "/cgi-bin/cdr"
    TIER = Tier()
    WEBSERVER = environ.get("SERVER_NAME") or TIER.hosts.get("APPC")
    DAY_ONE = URDATE
    PAGE_TITLE = "CDR Administration"
    TITLE = PAGE_TITLE
    SUBTITLE = None
    SUBMIT = "Submit"
    LOG_OUT = "Log Out"
    FORMATS = "html", "excel"
    LOGNAME = "reports"
    LOGLEVEL = "INFO"
    METHOD = "post"
    AUDIENCES = "Health Professional", "Patient"
    LANGUAGES = "English", "Spanish"
    INCLUDE_ANY_LANGUAGE_CHECKBOX = INCLUDE_ANY_AUDIENCE_CHECKBOX = False
    SUMMARY_SELECTION_METHODS = "id", "title", "board"
    EMAIL_PATTERN = re_compile(r"[^@]+@[^@\.]+\.[^@]+$")
    KEEP_COMPLETE_TITLES = False
    SAME_WINDOW = "jQuery('#primary-form').attr('target', '_self');"
    NONBREAKING_HYPHEN = "\u2011"

    def __init__(self, **opts):
        """Set up a skeletal controller."""

        self.__started = datetime.now()
        self.__opts = opts
        self.logger.info("started %s", self.subtitle or "controller")

    # ----------------------------------------------------------------
    # Top-level processing routines.
    # ----------------------------------------------------------------

    def run(self):
        """Override in derived class if there are custom actions."""

        try:
            return self.show_report() if self.request else self.show_form()
        except Exception as e:
            self.logger.exception("Controller.run() failure")
            self.bail(e)

    def show_form(self):
        """Populate an HTML page with a form and fields and send it."""

        self.populate_form(self.form_page)
        for label in self.buttons:
            button = self.form_page.button(label)
            if self.same_window and label in self.same_window:
                button.set("onclick", self.SAME_WINDOW)
            self.form_page.form.append(button)
        for alert in self.alerts:
            message = alert["message"]
            del alert["message"]
            self.form_page.add_alert(message, **alert)
        self.form_page.send()

    def show_report(self):
        """Override this method if you need to make some tweaks.

        A typical case would be:

            def show_report(self):
                self.report.body.set("id", "activity-report")
                self.report.send(self.format)

        Another situation in which you would override this method
        would be if you have a non-tabular report to be sent.
        """

        if self.format == "html":
            if self.use_basic_web_page:
                report = BasicWebPage()
                report.wrapper.append(report.B.H1(self.subtitle))
                tables = self.build_tables() or []
                if not isinstance(tables, (list, tuple)):
                    tables = [tables]
                for table in tables:
                    report.wrapper.append(table.node)
                report.wrapper.append(self.footer)
                return report.send()
            if self.report_css:
                self.report.page.add_css(self.report_css)
            elapsed = self.report.page.html.get_element_by_id("elapsed", None)
            if elapsed is not None:
                elapsed.text = str(self.elapsed)
            for alert in self.alerts:
                message = alert["message"]
                del alert["message"]
                self.report.page.add_alert(message, **alert)
            if self.wide_css:
                sibling = self.report.page.form.getparent()
                for table in self.report.page.form.findall("table"):
                    sibling.addnext(table)
                    sibling = table
                self.report.page.add_css(self.wide_css)
        self.report.send(self.format)

    def populate_form(self, page):
        """Stub, to be overridden by real controllers."""

    def build_tables(self):
        """Stub, to be overridden by real controllers."""
        return []

    # ----------------------------------------------------------------
    # General-purpose utility methods.
    # ----------------------------------------------------------------

    def load_group(self, group):
        """Fetch the active members of a named user group.

        Pass:
            group - name of group to fetch

        Return:
            dictionary-like object of user names indexed by user ID
        """

        query = db.Query("usr u", "u.id", "u.fullname", "u.name")
        query.join("grp_usr j", "j.usr = u.id")
        query.join("grp g", "g.id = j.grp")
        query.where("u.expired IS NULL")
        query.where(query.Condition("g.name", group))
        rows = query.execute(self.cursor).fetchall()

        class Group:
            def __init__(self, rows):
                self.map = {}
                for row in rows:
                    self.map[row.id] = row.fullname or row.name
                items = self.map.items()
                values = [(val[1].lower(), val[0], val[1]) for val in items]
                self.items = [vals[1:] for vals in sorted(values)]

            def __getvalue__(self, key):
                return self.map.get(key)

        return Group(rows)

    def load_valid_values(self, table_name):
        """Factor out logic for collecting a valid values set.

        This works because our tables for valid values have the
        same structure.

        Pass:
            table_name - name of the database table for the values

        Return:
            a populated `Values` object
        """

        query = self.Query(table_name, "value_id", "value_name")
        rows = query.order("value_pos").execute(self.cursor).fetchall()

        class Values:
            def __init__(self, rows):
                self.map = {}
                self.values = []
                for value_id, value_name in rows:
                    self.map[value_id] = value_name
                    self.values.append((value_id, value_name))
        return Values(rows)

    def log_elapsed(self):
        """Record how long this took."""
        self.logger.info(f"elapsed: {self.elapsed.total_seconds():f}")

    def make_url(self, script, **params):
        """Create a URL.

        Pass:
            script - string for base of url (can be relative or absolute)
            params - dictionary of named parameters for the URL

        Return:
            value appropriate for the href attribute of a link
        """

        if self.SESSION not in params:
            params[self.SESSION] = self.session.name
        params = urllib_parse.urlencode(params, doseq=True)
        return f"{script}?{params}"

    def redirect(self, where, session=None, **params):
        """Send the user to another page.

        Pass:
            where - base URL, up to but not including parameters
            session - session string or object to override this session (opt)
            params - dictionary of other named parameters
        """

        session = session or self.session
        self.navigate_to(where, session, **params)

    # ----------------------------------------------------------------
    # Routines specific to Summary reports.
    # ----------------------------------------------------------------

    def add_audience_fieldset(self, page):
        """Add radio buttons for PDQ audience.

        Pass:
            page - object on which we place the fields
        """

        fieldset = page.fieldset("Audience")
        fieldset.set("class", "by-board-block usa-fieldset")
        fieldset.set("id", "audience-block")
        default = self.default_audience
        if self.INCLUDE_ANY_AUDIENCE_CHECKBOX:
            checked = False if default else True
            opts = dict(label="Any", value="", checked=checked)
            fieldset.append(page.radio_button("audience", **opts))
        elif not default:
            default = self.AUDIENCES[0]
        for value in self.AUDIENCES:
            checked = True if value == default else False
            opts = dict(value=value, checked=checked)
            fieldset.append(page.radio_button("audience", **opts))
        page.form.append(fieldset)

    def add_board_fieldset(self, page):
        """Add checkboxes for the PDQ Editorial Boards.

        Pass:
            page - object on which we place the fields
        """

        fieldset = page.fieldset("Board")
        fieldset.set("class", "by-board-block usa-fieldset")
        fieldset.set("id", "board-set")
        boards = ["all"]
        if hasattr(self, "board") and isinstance(self.board, (list, tuple)):
            boards = self.board
        checked = "all" in boards or not boards
        opts = dict(label="All Boards", value="all", checked=checked)
        fieldset.append(page.checkbox("board", **opts))
        for value, label in self.get_boards().items():
            opts = dict(value=value, label=label, classes="ind")
            if value in boards:
                opts["checked"] = True
            fieldset.append(page.checkbox("board", **opts))
        page.form.append(fieldset)

    def add_language_fieldset(self, page):
        """Add radio buttons for summary language.

        Pass:
            page - object on which we place the fields
        """

        fieldset = page.fieldset("Language")
        fieldset.set("class", "by-board-block usa-fieldset")
        fieldset.set("id", "language-block")
        current = self.language if hasattr(self, "language") else None
        if self.INCLUDE_ANY_LANGUAGE_CHECKBOX:
            checked = not current
            opts = dict(label="Any", value="", checked=checked)
            fieldset.append(page.radio_button("language", **opts))
        elif not current:
            current = self.LANGUAGES[0]
        for value in self.LANGUAGES:
            checked = value == current
            opts = dict(value=value, checked=checked)
            fieldset.append(page.radio_button("language", **opts))
        page.form.append(fieldset)

    def add_summary_selection_fields(self, page, **kwopts):
        """
        Display the fields used to specify which summaries should be
        selected for a report, using one of several methods:

            * by summary document ID
            * by summary title
            * by summary board

        There are two branches taken by this method. If the user has
        elected to select a summary by summary title, and the summary
        title fragment matches more than one summary, then a follow-up
        page is presented on which the user selects one of the summaries
        and re-submits the report request. Otherwise, the user is shown
        options for choosing a selection method, which in turn displays
        the fields appropriate to that method dynamically. We also add
        JavaScript functions to handle the dynamic control of field display.

        Pass:
            page     - Page object on which to show the fields
            titles   - an optional array of SummaryTitle objects
            audience - if False, omit Audience buttons (default is True)
            language - if False, omit Language buttons (default is True)
            id-label - optional string for the CDR ID field (defaults
                       to "CDR ID" but can be overridden, for example,
                       to say "CDR ID(s)" if multiple IDs are accepted)
            id-tip   - optional string for the CDR ID field for popup
                       help (e.g., "separate multiple IDs by spaces")

        Return:
            nothing (the form object is populated as a side effect)
        """

        # --------------------------------------------------------------
        # Show the second stage in a cascading sequence of the form if we
        # have invoked this method directly from build_tables(). Widen
        # the form to accomodate the length of the title substrings
        # we're showing.
        # --------------------------------------------------------------
        titles = kwopts.get("titles")
        if titles:
            page.form.append(page.hidden_field("selection_method", "id"))
            fieldset = page.fieldset("Choose Summary")
            page.add_css("fieldset { width: 600px; }")
            for t in titles:
                opts = dict(label=t.display, value=t.id, tooltip=t.tooltip)
                fieldset.append(page.radio_button("cdr-id", **opts))
            page.form.append(fieldset)

        else:

            # Fields for the original form.
            fieldset = page.fieldset("Selection Method")
            methods = "PDQ Board", "CDR ID", "Summary Title"
            for method in methods:
                value = method.split()[-1].lower()
                checked = value == self.selection_method
                opts = dict(label=f"By {method}", value=value, checked=checked)
                fieldset.append(page.radio_button("selection_method", **opts))
            page.form.append(fieldset)
            self.add_board_fieldset(page)
            if kwopts.get("audience", True):
                self.add_audience_fieldset(page)
            if kwopts.get("language", True):
                self.add_language_fieldset(page)
            fieldset = page.fieldset("Summary Document ID")
            fieldset.set("class", "by-id-block usa-fieldset")
            label = kwopts.get("id-label", "CDR ID")
            opts = dict(label=label, tooltip=kwopts.get("id-tip"))
            if hasattr(self, "cdr_id"):
                if isinstance(self.cdr_id, (int, str)):
                    opts["value"] = self.cdr_id
            fieldset.append(page.text_field("cdr-id", **opts))
            page.form.append(fieldset)
            fieldset = page.fieldset("Summary Title")
            fieldset.set("class", "by-title-block usa-fieldset")
            opts = dict(tooltip="Use wildcard (%) as appropriate.")
            if hasattr(self, "fragment") and self.fragment:
                opts["value"] = self.fragment
            fieldset.append(page.text_field("title", **opts))
            page.form.append(fieldset)
            page.add_script(self.summary_selection_js)

    def get_boards(self):
        """Construct a dictionary of PDQ board names indexed by CDR ID."""

        boards = Board.get_boards().values()
        OD = OrderedDict
        return OD([(board.id, board.short_name) for board in boards])

    # ----------------------------------------------------------------
    # Static and class methods.
    # ----------------------------------------------------------------
    @staticmethod
    def add_date_range_to_caption(caption, start, end):
        """Format caption with date range (we do this a lot).

        Pass:
            caption - base string for the start of the caption
            start - optional beginning of the date range
            end - optional finish of the date range

        Return:
            possibly altered string for the table caption
        """

        if start:
            if end:
                return f"{caption} Between {start} and {end}"
            return f"{caption} Since {start}"
        elif end:
            return f"{caption} Through {end}"
        return caption

    @classmethod
    def bail(cls, message=TAMPERING, /, **opts):
        """Send an error page to the browser.

        Optional positional argument:
          message - string describing the problem
                    by default this is a vague intended to convey as little
                    information to a hacker as possible

        Optional keyword arguments:
          extra - sequence of extra lines to append
          logfile - name of logfile to write to
        """

        class ErrorPage(HTMLPage):
            def __init__(self, message, extra):
                HTMLPage.__init__(self, "CDR Error")
                self.message = message
                self.extra = extra
                if extra and not isinstance(extra, (list, tuple)):
                    self.extra = [extra]

            @cached_property
            def main(self):
                alert_body = self.B.DIV(
                    self.B.H3(
                        str(self.message),
                        self.B.CLASS("usa-alert__heading")
                    )
                )
                if self.extra:
                    if len(self.extra) == 1:
                        p = self.B.P(str(self.extra[0]))
                        p.set("class", "usa-alert__text")
                        alert_body.append(p)
                    else:
                        extra = self.B.UL()
                        for arg in self.extra:
                            extra.append(self.B.LI(str(arg)))
                        alert_body.append(extra)
                alert_body.set("class", "usa-alert__body")
                return self.B.E(
                    "main",
                    self.B.DIV(
                        self.B.E(
                            "section",
                            self.B.DIV(
                                alert_body,
                                self.B.CLASS("usa-alert usa-alert--error")
                            )
                        ),
                        self.B.CLASS("grid-container")
                    ),
                    self.B.CLASS("usa-section")
                )

        try:
            page = ErrorPage(message, opts.get("extra"))
        except Exception:
            page = ErrorPage(cls.TAMPERING, {})
        logfile = opts.get("logfile")
        if logfile:
            if logfile.lower().endswith(".log"):
                logfile = logfile[:-4]
            logger = Logging.get_logger(logfile)
            logger.error("cdrcgi bailout: %s", message)
        page.send()

    @classmethod
    def navigate_to(cls, where, session, **params):
        """Send the user to another page.

        This is the non-instance version.
        Pass:
            where - base URL, up to but not including parameters (required)
            session - session string or object (required)
            params - dictionary of other named parameters
        """

        where = where.split("?")[0]
        params[__class__.SESSION] = session
        params = urllib_parse.urlencode(params)
        print(f"Location:https://{cls.WEBSERVER}{cls.BASE}/{where}?{params}\n")
        sys_exit(0)

    @staticmethod
    def parse_date(iso_date):
        """Convert a date string to a `datetime.date` object.

        Changed requirements: might not be an ISO date any more, because
        for some strange reasone, the USWDS project uses a non-standard
        format.

        Pass:
            iso_date - optional string for the date

        Return:
            None if the string is None or empty, otherwise a date object
        """

        if iso_date is None or not iso_date.strip():
            return None
        if "/" in iso_date:
            month, day, year = iso_date.strip().split("/")
        else:
            year, month, day = iso_date.strip().split("-")
        return date(int(year), int(month), int(day))

    @classmethod
    def parse_email_address(cls, address):
        """Pull out an email address from a string.

        Performs a very simple validation, may improve it later, but full
        RFC requires an incredible thousand-character regular expression.

        Pass:
            address - string which might have a display portion
                      (e.g., "Joe Blow <joe@example.com>")

        Return:
            address portion of the string if validation passes
            None if no valid address is found
        """

        realname, address = parse_email_address(address)
        if address and cls.EMAIL_PATTERN.match(address):
            if ".." not in address:
                return address
        return None

    @staticmethod
    def send_page(page, text_type="html", mime_type=None):
        """Send a string back to the web server using UTF-8 encoding.

        Required position argument:
            page - Unicode string for the page or DOM object

        Optional keyword arguments:
            text_type - typically "html" but sometimes "xml"
            mime_type - for other types; for example, "application/json"
        """

        if not isinstance(page, str):
            opts = dict(HTMLPage.STRING_OPTS, encoding="unicode")
            page = lxml_html.tostring(page, **opts)
        mime_type = mime_type or f"text/{text_type}"
        string = f"Content-type: {mime_type};charset=utf-8\n\n{page}"
        sys_stdout.buffer.write(string.encode("utf-8"))
        sys_exit(0)

    @staticmethod
    def toggle_display(function_name, show_value, class_name):
        """Create JavaScript function to show or hide elements.

        Pass:
            function_name  - name of the JavaScript function to create
            show_value     - controlling element's value causing show
            class_name     - class of which the controlled blocks are members
        Return:
            source code for JavaScript function
        """

        return f"""\
function {function_name}(value) {{
    if (value == "{show_value}")
        jQuery(".{class_name}").show();
    else
        jQuery(".{class_name}").hide();
}}"""

    # ----------------------------------------------------------------
    # Instance properties.
    # ----------------------------------------------------------------

    @cached_property
    def alerts(self):
        """Override to add alerts which should be shown on the page."""
        return []

    @cached_property
    def buttons(self):
        """Sequence of names for request buttons to be provided."""

        buttons = self.__opts.get("buttons")
        if buttons is None:
            if self.SUBMIT:
                return [self.SUBMIT]
            return []
        return buttons

    @cached_property
    def conn(self):
        """Database connection for this controller."""
        return db.connect()

    @cached_property
    def cursor(self):
        """Database cursor for this controller."""
        return self.conn.cursor()

    @property
    def default_audience(self):
        """Let a subclass override the default for the audience picklist."""
        return None

    @cached_property
    def doc_titles(self):
        """Cached lookup of CDR document titles by ID.

        By default, only the portion of the title column's value before
        the first semicolon is used. If "Inactive;" is at the front of
        the title string the second segment of the title is used instead
        (if it exists) and " (inactive)" is appended. To preserve the
        entire contents of the title column's values, set the class-level
        property `KEEP_COMPLETE_TITLES` to `True` in the derived class.
        """

        class DocTitles(UserDict):

            def __init__(self, control):
                self.__control = control
                UserDict.__init__(self)

            def __getitem__(self, key):
                if key not in self.data:
                    query = self.__control.Query("document", "title")
                    query.where(query.Condition("id", key))
                    row = query.execute(self.__control.cursor).fetchone()
                    title = row.title.strip() if row else ""
                    if not self.__control.KEEP_COMPLETE_TITLES:
                        pieces = [p.strip() for p in row.title.split(";")]
                        title = pieces[0]
                        if title.lower() == "inactive" and len(pieces) > 1:
                            title = f"{pieces[1]} (inactive)"
                    self.data[key] = title or None
                return self.data[key]

        return DocTitles(self)

    @property
    def elapsed(self):
        """How long have we been running? Don't cache."""
        return datetime.now() - self.started

    @cached_property
    def fields(self):
        """CGI fields for the web form."""
        return FieldStorage()

    @cached_property
    def footer(self):
        """Override to alter or suppress the default report footer."""

        user = self.session.User(self.session, id=self.session.user_id)
        name = user.fullname or user.name
        today = date.today()
        generated = f"Report generated {today} by {name}"
        elapsed = HTMLPage.B.SPAN(str(self.elapsed), id="elapsed")
        args = generated, HTMLPage.B.BR(), "Elapsed: ", elapsed
        footer = HTMLPage.B.P(*args)
        footer.set("class", "report-footer")
        return footer

    @cached_property
    def format(self):
        """Either "html" (the default) or "excel"."""

        format = self.fields.getvalue("format")
        if not format:
            format = self.__opts.get("format") or self.FORMATS[0]
        if format not in self.FORMATS:
            self.bail("invalid report format")
        return format

    @cached_property
    def form_page(self):
        """Create a form page."""

        opts = {
            "control": self,
            "action": self.script,
            "buttons": [HTMLPage.button(b) for b in self.buttons],
            "subtitle": self.subtitle,
            "session": self.session,
            "method": self.method,
            "suppress_sidenav": self.suppress_sidenav,
        }
        return self.HTMLPage(self.title, **opts)

    @cached_property
    def HTMLPage(self):
        """Allow overriding of page class."""
        return HTMLPage

    @cached_property
    def logged_out(self):
        """True if the user has just logged out."""
        return True if self.fields.getvalue("logged_out") else False

    @cached_property
    def logger(self):
        """Object for recording what we do."""

        logger = self.__opts.get("logger")
        if logger is not None:
            return logger
        opts = dict(level=self.loglevel)
        return Logging.get_logger(self.LOGNAME, **opts)

    @cached_property
    def loglevel(self):
        """Override this to provide runtime control of logging."""
        return self.LOGLEVEL

    @cached_property
    def method(self):
        """Allow override of form method."""
        return self.fields.getvalue("method") or self.METHOD

    @cached_property
    def no_results(self):
        """Message to display if no result tables are returned."""
        return "No report results found."

    @cached_property
    def Query(self):
        """Convenience reference to database query class object."""
        return db.Query

    @cached_property
    def report(self):
        """Create the `Reporter` object for this job."""

        tables = self.build_tables()
        if self.format == "excel":
            return Reporter(self.title, tables)
        page_opts = dict(
            session=self.session,
            action=self.script or None,
            control=self,
        )
        opts = dict(
            footer=self.footer,
            subtitle=self.subtitle,
            no_results=self.no_results,
            page_opts=page_opts,
        )
        return Reporter(self.title, tables, **opts)

    @cached_property
    def report_css(self):
        """Override to provide additional styling to a reports page."""
        return None

    @cached_property
    def Reporter(self):
        """Allow overriding of page class."""
        return Reporter

    @cached_property
    def request(self):
        """Name of clicked request button, if any."""

        request = self.fields.getvalue(self.REQUEST, "").strip()
        if re_search("[^A-Za-z0-9 -]", request):
            self.logger.warning("bad request %r", request)
            return self.bail()
        return request

    @cached_property
    def same_window(self):
        """Override for commands which should stay in the same window"""
        return []

    @cached_property
    def script(self):
        """Name of form submission handler."""
        return self.__opts.get("script") or Path(sys_argv[0]).name

    @cached_property
    def selection_method(self):
        """How does the user want to identify summaries for the report?"""

        method = self.fields.getvalue("selection_method", "board")
        if method not in self.SUMMARY_SELECTION_METHODS:
            self.bail()
        return method

    @cached_property
    def session(self):
        """Session object for this controller.

        Note: this is an object, not a string. For the session name,
        use `self.session.name` or `str(self.session)` or `f"{self.session}"`.

        No need to specify a tier here, as web CGI scripts are only
        intended to work on the local tier.
        """

        session = self.__opts.get("session")
        if not session:
            session = self.fields.getvalue(self.SESSION) or "guest"
        if isinstance(session, (list, tuple)):
            session = session[0]
        if isinstance(session, bytes):
            session = str(session, "ascii")
        if isinstance(session, str):
            try:
                session = Session(session)
            except Exception as e:
                self.bail("Invalid or expired session.")
        if not isinstance(session, Session):
            raise Exception(f"{session}: Not a session object")
        return session

    @cached_property
    def show_news(self):
        """Whether we should display news announcements."""
        return True if self.fields.getvalue("show_news") else False

    @cached_property
    def summary_selection_js(self):
        "Local JavaScript to manage sections of the form dynamically."

        return """\
function check_set(name, val) {
    var all_selector = "#" + name + "-all";
    var ind_selector = "#" + name + "-set .ind";
    if (val == "all") {
        if (jQuery(all_selector).prop("checked"))
            jQuery(ind_selector).prop("checked", false);
        else
            jQuery(all_selector).prop("checked", true);
    }
    else if (jQuery(ind_selector + ":checked").length > 0)
        jQuery(all_selector).prop("checked", false);
    else
        jQuery(all_selector).prop("checked", true);
}
function check_board(board) { check_set("board", board); }
function check_selection_method(method) {
    switch (method) {
        case 'id':
            jQuery('.by-board-block').hide();
            jQuery('.by-id-block').show();
            jQuery('.by-title-block').hide();
            break;
        case 'board':
            jQuery('.by-board-block').show();
            jQuery('.by-id-block').hide();
            jQuery('.by-title-block').hide();
            break;
        case 'title':
            jQuery('.by-board-block').hide();
            jQuery('.by-id-block').hide();
            jQuery('.by-title-block').show();
            break;
    }
}
jQuery(function() {
    var method = jQuery("input[name='selection_method']:checked").val();
    check_selection_method(method);
});"""

    @cached_property
    def summary_titles(self):
        """Find the summaries that match the user's title fragment.

        Note that the user is responsible for adding any non-trailing
        SQL wildcards to the fragment string. If the title is longer
        than 60 characters, truncate with an ellipsis, but add a
        tooltip showing the whole title. We create a local class for
        the resulting list.

        ONLY WORKS IF YOU IMPLEMENT THE `self.fragment` PROPERTY!!!
        """

        titles = None
        if hasattr(self, "fragment") and self.fragment:
            class SummaryTitle:
                def __init__(self, doc_id, display, tooltip=None):
                    self.id = doc_id
                    self.display = display
                    self.tooltip = tooltip
            fragment = f"{self.fragment}%"
            query = self.Query("active_doc d", "d.id", "d.title")
            query.join("doc_type t", "t.id = d.doc_type")
            query.where("t.name = 'Summary'")
            query.where(query.Condition("d.title", fragment, "LIKE"))
            query.order("d.title")
            rows = query.execute(self.cursor).fetchall()
            titles = []
            for doc_id, title in rows:
                if len(title) > 60:
                    short_title = title[:57] + "..."
                    summary = SummaryTitle(doc_id, short_title, title)
                else:
                    summary = SummaryTitle(doc_id, title)
                titles.append(summary)
        return titles

    @cached_property
    def suppress_sidenav(self):
        """Override to implement more nuanced logic."""
        return False

    @cached_property
    def timestamp(self):
        """String used to distinguish multiple instances of named items."""
        return self.started.strftime("%Y%m%d%H%M%S")

    @cached_property
    def started(self):
        """When did we start processing?"""
        return self.__started

    @cached_property
    def subtitle(self):
        """String to be displayed as the page title (bad name)."""
        return self.__opts.get("subtitle") or self.SUBTITLE

    @cached_property
    def title(self):
        """String used as the browser title for the page."""
        return self.__opts.get("title") or self.TITLE or self.PAGE_TITLE

    @cached_property
    def use_basic_web_page(self):
        """True if report should use the alternate (simpler) layout."""
        return False

    @cached_property
    def wide_css(self):
        """Override to allow tables in a report more width."""
        return None


class FormFieldFactory:
    """Provide class methods for creating HTML form fields.

    Also has a factory method for creating a fieldset (with an optional
    legend) which isn't really a field, but it seemed to fit best here.

    These methods are inherited by the HTMLPage class below, so that
    users passed an object of that class have easy access to the
    methods without having to import anything extra.

    Here are the categories of fields supported:

        text fields
           - text_field()
           - password_field()
           - date_field()

        clickable fields
           - checkbox()
           - radio_button()

        other
           - file_field()
           - hidden_field()
           - select()

    Pro tip:
       If you don't specify a label for a field, the class will create
       one for you based on the field's value, by replacing underscores
       with spaces, and capitalizing each word. Of course, this is only
       useful for fields which have an initial value (as is always the
       case for radio buttons and checkboxes). For example:

       page.radio_button("sort", value="by_board")

       will result in a field whose label is "By Board".
    """

    EN_DASH = "\u2013"
    EM_DASH = "\u2014"
    LINK_COLOR = "#005ea2"
    CLICKABLE = "checkbox", "radio"
    B = html_builder

    @classmethod
    def accordion(cls, name, **kwargs):
        """Collapsable wrapper.

        Required positional argument:
          name - unique name for this wrapper

        Optional keyword arguments:
          label - text to display in the header
          open - start with the section uncollapsed
          prose - if True add class usa-prose to payload
          bordered - if True add a border to the wrapper

        Return:
          object with wrapper and payload properties
        """

        id = f"{name}-accordion"
        label = kwargs.get("label")
        if not label:
            label = name.replace("-", " ").replace("_", " ").title()
        button = cls.B.BUTTON(
            label,
            cls.B.CLASS("usa-accordion__button"),
            type="button",
        )
        payload_classes = ["usa-accordion__content"]
        wrapper_classes = ["usa-accordion"]
        if kwargs.get("prose"):
            payload_classes.append("usa-prose")
        if kwargs.get("bordered"):
            wrapper_classes.append("usa-accordion--bordered")
        button.set("aria-expanded", "true" if kwargs.get("open") else "false")
        button.set("aria-controls", id)
        accordion = SimpleNamespace()
        accordion.payload = cls.B.DIV(id=id)
        header = cls.B.H4(button)
        header.set("class", "usa-accordion__heading")
        accordion.payload.set("class", " ".join(payload_classes))
        accordion.wrapper = cls.B.DIV(header, accordion.payload)
        accordion.wrapper.set("class", " ".join(wrapper_classes))
        return accordion

    @classmethod
    def button(cls, label, **kwargs):
        """Create a button to be added to an HTMLPage object.

        These are called "buttons" even though the HTML element used
        to create the widget is an "input" element with type of
        "submit" or "reset" (and even though, confusingly, there is
        another HTML element called "button," which is not being used
        here).

        Not to be confused with radio buttons. :-)

        Required positional argument:

            label
                string to be displayed on the button; also used by the
                form submission handler to determine which button was
                clicked

        Optional keyword arguments:

            button_type
                set to "reset" to override the default of "submit"

            onclick
                optional JavaScript to be invoked when the button
                is clicked

        Return:

            lxml object for an INPUT element

        """

        id_label = label.lower().replace(" ", "-")
        button = cls.B.INPUT(value=label, name=Controller.REQUEST)
        button.set("type", kwargs.get("button_type", "submit"))
        button.set("class", "button usa-button")
        button.set("id", f"submit-button-{id_label}")
        onclick = kwargs.get("onclick")
        if onclick:
            button.set("onclick", onclick)
        return button

    @classmethod
    def checkbox(cls, group, **kwargs):
        """Create a checkbox field block with label.

        Required positional argument:

            group

                string identifying the set of checkboxes whose checked
                values will be returned to the form processor as a
                single sequence

        See the `__field()`, `__wrapper()`, and `__clickable()`
        methods for a description of the available optional keyword
        arguments.

        Example:

            fieldset = page.fieldset("Select Board(s)")
            for board in self.boards:
                opts = dict(value=board.id, label=board.name)
                fieldset.append(page.checkbox("board", **opts)
            page.append(fieldset)

        Return:

            lxml object for a wrapper element, enclosing an INPUT element
        """

        return cls.__clickable(group, "checkbox", **kwargs)

    @classmethod
    def date_field(cls, name, **kwargs):

        """Create a date field block with optional label.

        A date picker is displayed for ease of choosing a value for the
        field. The text widget is still editable directly.

        Required positional argument:

            name
                unique name of the field on the page; it is essential
                that the name be unique, even across forms (in the
                event that the caller adds a second form to the page)

        See the `__field()` and `__wrapper()` methods for a
        description of the available optional keyword arguments.

        Return:

            lxml object for a wrapper element, enclosing an INPUT element
        """

        field = cls.__field(name, "text", **kwargs)
        value = kwargs.get("value")
        label = kwargs.get("label", name.replace("_", " ").title())
        label = cls.B.LABEL(label, cls.B.FOR(field.get("id")))
        label.set("class", "usa-label")
        field.set("type", "date")
        date_picker = cls.B.DIV(field, cls.B.CLASS("usa-date-picker"))
        if value:
            date_picker.set("data-default-value", str(value))
        return cls.B.DIV(
            label,
            date_picker,
            cls.B.CLASS("usa-form-group")
        )

    @classmethod
    def date_range(cls, name, **kwargs):
        """Date an inline pair of dates with a single label.

        Required positional argument:

            name
                unique name of the field on the page; it is essential
                that the name be unique, even across forms (in the
                event that the caller adds a second form to the page)

        Optional keyword arguments:

            start_date
                initial value for the first date field

            end_date
                initial value for the second date field

        See the `__field()` and `__wrapper()` methods for a
        description of other available optional keyword arguments.

        Return:

            lxml object for a wrapper element, enclosing two INPUT elements
            with a separating SPAN between
        """

        opts = dict(**kwargs)
        wrapper = cls.__wrapper(name, **opts)
        inner = cls.B.DIV(cls.B.CLASS("date-range-fields"))
        wrapper.append(inner)
        for which in ("start", "end"):
            opts["value"] = value = opts.get(f"{which}_date")
            field = cls.__field(f"{name}-{which}", "text", **opts)
            field.set("type", "date")
            date_picker = cls.B.DIV(field)
            date_picker.set("class", "usa-date-picker xxmargin-top-0")
            if value:
                date_picker.set("data-default-value", str(value))
            group = cls.B.DIV(date_picker)
            group.set("class", "usa-form-group margin-top-0")
            inner.append(group)
            if which == "start":
                separator = cls.B.SPAN(cls.EM_DASH)
                separator.set("class", "date-range-sep margin-right-2")
                inner.append(separator)
        return wrapper

    @classmethod
    def fieldset(cls, legend=None, **opts):
        """Create an HTML fieldset element with an optional legend child.

        Optional keyword argument:

            legend

                string for optional legend to be displayed for the
                fieldset

            id

                unique ID for the element

        Return:

            lxml object for a FIELDSET element
        """

        fieldset = cls.B.FIELDSET(cls.B.CLASS("usa-fieldset"))
        if legend:
            classes = "usa-legend"
            fieldset.append(cls.B.LEGEND(legend, cls.B.CLASS(classes)))
        if opts.get("id"):
            fieldset.set("id", opts.get("id"))
        return fieldset

    @classmethod
    def file_field(cls, name, **kwargs):
        """Create a file upload field block with optional label.

        Required positional argument:

            name
                unique name of the field on the page; it is essential
                that the name be unique, even across forms (in the
                event that the caller adds a second form to the page)

        See the `__field()` and `__wrapper()` methods for a
        description of the available optional keyword arguments.

        Return:

            lxml object for a wrapper element, enclosing an INPUT element
        """

        field = cls.__field(name, "file", **kwargs)
        if kwargs.get("multiple"):
            field.set("multiple")
        wrapper = cls.__wrapper(name, **kwargs)
        wrapper.append(field)
        return wrapper

    @classmethod
    def hidden_field(cls, name, value):
        if value is None:
            value = ""
        return cls.B.INPUT(name=name, value=str(value), type="hidden")

    @classmethod
    def password_field(cls, name, **kwargs):
        """Create a password field block with optional label.

        The value for the field is not displayed.

        Required positional argument:

            name
                unique name of the field on the page; it is essential
                that the name be unique, even across forms (in the
                event that the caller adds a second form to the page)

        See the `__field()` and `__wrapper()` methods for a
        description of the available optional keyword arguments.

        Return:

            lxml object for a wrapper element, enclosing an INPUT element
        """

        field = cls.__field(name, "password", **kwargs)
        wrapper = cls.__wrapper(name, **kwargs)
        wrapper.append(field)
        return wrapper

    @classmethod
    def radio_button(cls, group, **kwargs):
        """Create a wrapped radio button field with label.

        Required positional argument:

            group

                string identifying the set of radio buttons of which
                at most only one can be checked

        See the `__field()`, `__wrapper()`, and `__clickable()`
        methods for a description of the available optional keyword
        arguments.

        Example:

            choices = ("html", "Web Page"), ("excel", "Excel Workbook")
            fieldset = page.fieldset("Report Format")
            page.append(fieldset)
            for value, label in choices:
                opts = dict(label=label, value=value)
                if value == "html":
                    opts["checked"] = True
                fieldset.append(page.radio_button("format", **opts))

        Return:

            lxml object for a wrapper element, enclosing an INPUT element
        """

        return cls.__clickable(group, "radio", **kwargs)

    @classmethod
    def select(cls, name, **kwargs):
        """Create a wrapped picklist field with optional label.

        Required positional argument:

            name
                unique name of the field on the page; it is essential
                that the name be unique, even across forms (in the
                event that the caller adds a second form to the page)

        Optional keyword arguments:

            options
                sequence of choices

            size
                height of the select box (default 1; 4 if multiple)

            onchange
                javascript to execute when the selection changes

            multiple
                whether multiple selections are permitted

        See the `__field()` and `__wrapper()` methods for a description
        of additional available optional keyword arguments.

        Return:

            lxml object for a wrapper element, enclosing an SELECT element
        """

        # Create the widget and its wrapper element.
        wrapper = cls.__wrapper(name, **kwargs)
        field = cls.__field(name, "select", **kwargs)
        wrapper.append(field)

        # Add attributes unique to picklist fields.
        multiple = True if kwargs.get("multiple") else False
        if multiple:
            field.set("multiple")
        if kwargs.get("onchange"):
            field.set("onchange", kwargs["onchange"])
        if kwargs.get("size"):
            field.set("size", str(kwargs["size"]))

        # If we have any options, add them as children of the widget object.
        options = kwargs.get("options")
        if options:
            default = kwargs.get("default")
            if not isinstance(default, (list, tuple, set)):
                default = [default] if default else []
            default = {str("" if d is None else d) for d in default}
            if not multiple and len(default) > 1:
                error = "Multiple defaults specified for single picklist"
                raise Exception(error)
            if isinstance(options, dict):
                options = sorted(options.items(),
                                 key=lambda o: str(o[1]).lower())
            for option in options:
                if isinstance(option, (list, tuple)):
                    value, display = option
                else:
                    value = display = option
                if value is None:
                    value = ""
                option = cls.B.OPTION(str(display), value=str(value))
                if str(value) in default:
                    option.set("selected")
                field.append(option)

        # Done.
        return wrapper

    @classmethod
    def textarea(cls, name, **kwargs):
        """Create a textarea field block with optional label.

        Required positional argument:

            name
                unique name of the field on the page; it is essential
                that the name be unique, even across forms (in the
                event that the caller adds a second form to the page)

        See the `__field()` and `__wrapper()` methods for a
        description of the available optional keyword arguments.

        Return:

            lxml object for a wrapper element, enclosing an INPUT element
        """

        field = cls.__field(name, "textarea", **kwargs)
        if "rows" in kwargs:
            field.set("rows", str(kwargs["rows"]))
        wrapper = cls.__wrapper(name, **kwargs)
        wrapper.append(field)
        return wrapper

    @classmethod
    def text_field(cls, name, **kwargs):
        """Create a text field block with optional label.

        Required positional argument:

            name
                unique name of the field on the page; it is essential
                that the name be unique, even across forms (in the
                event that the caller adds a second form to the page)

        See the `__field()` and `__wrapper()` methods for a
        description of the available optional keyword arguments.

        Return:

            lxml object for a wrapper element, enclosing an INPUT element
        """

        field = cls.__field(name, "text", **kwargs)
        wrapper = cls.__wrapper(name, **kwargs)
        wrapper.append(field)
        return wrapper

    @classmethod
    def __classes(cls, classes):
        """Package the classes to be assigned to an element as a set

        Required positional argument:

            classes
                one of the following:
                   - None
                   - a space-delimited string of class names
                   - a `tuple` of class names
                   - a `list` of class names
                   - a `set` of class names

        Returns:

            a `set` of zero or more class name strings
        """

        if classes is None:
            return set()
        if isinstance(classes, str):
            return set(classes.split())
        return set(classes)

    @classmethod
    def __clickable(cls, group, field_type, **kwargs):
        """Create a wrapped checkbox or radio button field.

        Required positional arguments:

            group
                string identifying a set of related clickable fields
                which will be returned as a sequence of checked
                values to the form processor

            field_type
                one of "checkbox" or "radio"

        See the `__field()` and `__wrapper()` methods for a
        description of the available optional keyword arguments.

        Return:

            lxml object for a wrapper element, enclosing an INPUT element
        """

        # Create the field and its wrapper.
        widget = cls.__field(group, field_type, **kwargs)
        widget.tail = " "
        kwargs["field_type"] = field_type
        wrapper = cls.__wrapper(group, clickable=True, **kwargs)

        # Add attributes unique to radio buttons and checkboxes.
        if kwargs.get("checked"):
            widget.set("checked")
        value = kwargs.get("value")
        if value is None:
            value = ""
        value = str(value)
        onclick = kwargs.get("onclick", f"check_{group}('{value}')")
        if onclick:
            widget.set("onclick", onclick.replace("-", "_"))

        # For these fields, the label follows the widget.
        label = kwargs.get("label", value.replace("_", " ").title())
        label = cls.B.LABEL(cls.B.FOR(widget.get("id")), label)
        label_class = f"usa-{field_type}__label"
        label.set("class", f"clickable {label_class}")
        if kwargs.get("tooltip"):
            label.set("title", kwargs["tooltip"])
        wrapper.append(widget)
        wrapper.append(label)
        return wrapper

    @classmethod
    def __field(cls, name, field_type, **kwargs):
        """Build an HTML form field widget.

        Required positional arguments:

            name
                unique name of the field on the page (not just unique
                to the form); be careful also to avoid clashes with
                names generated for date range fields ({name}-start
                and {name}-end)

            field_type
                one of the following string values:
                   - text
                   - password
                   - file
                   - textarea
                   - select
                   - checkbox
                   - radio

        Optional keyword arguments:

            value
                initial value for the field (default: "")
                not used for "select" fields

            widget_id
                custom value for the field element's ID attribute
                (default: normalized name-value string for radio
                buttons and checkboxes, name for all other field
                types)

            classes
                if present, used as the 'class' attribute for the
                field element; may be passed as a list, tuple, or set,
                or as a string value with one or more space-delimited
                class names (default: no classes)

            tooltip
                if present, used as the 'title' attribute for the
                field element (default: no tooltip)

            readonly
                if True, the field is not editable, but it can get the
                UI focus, and it is still sent with the form submission
                (default: False); not used for 'select' fields

            disabled
                like readonly, but the form element does not receive
                focus, is skipped in tabbing navigation, and is not
                passed to the form processor

        Returns:

            lxml HTML element object
        """

        # Create the object for the field element.
        value = kwargs.get("value")
        if value is None:
            value = ""
        value = str(value)
        if field_type == "textarea":
            field = cls.B.TEXTAREA(value, name=name)
        elif field_type == "select":
            field = cls.B.SELECT(name=name)
        else:
            field = cls.B.INPUT(name=name, value=value)
            field.set("type", field_type)

        # Add the element's unique ID attribute.
        widget_id = kwargs.get("widget_id")
        if not widget_id:
            if field_type in cls.CLICKABLE:
                widget_id = f"{name}-{value}".replace(" ", "-").lower()
            else:
                widget_id = name
        field.set("id", widget_id)

        # Add the classes for the widget element.
        classes = cls.__classes(kwargs.get("classes"))
        if field_type == "textarea" and "usa-textarea" not in classes:
            classes.add("usa-textarea")
        elif field_type in ("text", "password") and "usa-input" not in classes:
            classes.add("usa-input")
            classes.add("usa-input--xl")
        elif field_type == "select":
            classes.add("usa-select")
            classes.add("form-select")
        elif field_type == "checkbox":
            classes.add("usa-checkbox__input")
        elif field_type == "radio":
            classes.add("usa-radio__input")
        elif field_type == "file":
            classes.add("usa-file-input")
        if classes:
            field.set("class", " ".join(classes))

        # Tooltip for clickables goes on the label, not the widget.
        elif kwargs.get("tooltip"):
            field.set("title", kwargs["tooltip"])

        # Add the attributes which (almost) any field can use, and we're done.
        if kwargs.get("disabled"):
            field.set("disabled")
        if kwargs.get("readonly") and field_type != "select":
            field.set("readonly")
        return field

    @classmethod
    def __wrapper(cls, name, **kwargs):
        """Create an enclosing element for a form field

        Required positional argument:

            name
                unique name for the field

        Optional keyword arguments:

            clickable
                True if the field is a checkbox or radio button

            field_type
                For example, "checkbox"

            wrapper
                string value for the name of the element to use
                for the field's enclosing element (default: "div")

            label
                string for identify which field this is

            wrapper_id
                custom value for the enclosing element's ID attribute
                (default: None)

            wrapper_classes
                one or more classes to be applied to the div
                wrapper (as space-delimited string, or as list,
                tuple, or set) or None if no additional classes
                are to be applied; the "labeled-field" class
                will be added to the set of classes for fields
                labeled on the left

        Returns:

            lxml HTML element object
        """

        # Create the enclosing element.
        wrapper = cls.B.E(kwargs.get("wrapper") or "div")

        # Set the unique ID for the enclosing element.
        if kwargs.get("wrapper_id"):
            wrapper.set("id", kwargs["wrapper_id"])

        # Determine the classes to be applied to the wrapper element.
        classes = cls.__classes(kwargs.get("wrapper_classes"))
        field_type = kwargs.get("field_type")
        if field_type == "checkbox":
            classes.add("usa-checkbox")
        elif field_type == "radio":
            classes.add("usa-radio")
        elif field_type == "file":
            classes.add("usa-form-group")
        label = kwargs.get("label", name.replace("_", " ").title())
        if label and not kwargs.get("clickable"):
            classes.add("labeled-field")
            widget_id = kwargs.get("widget_id") or name
            label = cls.B.LABEL(label, cls.B.FOR(widget_id))
            label.set("class", "usa-label")
            if kwargs.get("tooltip"):
                label.set("title", kwargs["tooltip"])
            wrapper.append(label)
        if classes:
            wrapper.set("class", " ".join(classes))

        # Done.
        return wrapper


class HTMLPage(FormFieldFactory):
    """Web page for the CDR system.

    Replacement for the old `Page` class, which built up a page as a
    sequence of strings. This class does it the way it should always
    have been done, by building an object tree, using the html.builder
    support in the lxml package.

    Be sure to keep the JQUERY... class values. They're not used here,
    but other scripts use them.
    """

    BASE = Controller.BASE
    VERSION = "202101071440"
    APIS = "https://ajax.googleapis.com/ajax/libs"
    JQUERY = "/js/jquery.min.js"
    JQUERY_UI = "/js/jquery-ui.min.js"
    JQUERY_CSS = "/stylesheets/jquery-ui.min.css"
    CSS_LINKS = (
        dict(href="/uswds/css/uswds.min.css", rel="stylesheet"),
        dict(href=JQUERY_CSS, rel="stylesheet"),
    )
    SCRIPT_LINKS = dict(src=JQUERY), dict(src=JQUERY_UI)
    PRIMARY_FORM_ID = "primary-form"
    STRING_OPTS = dict(pretty_print=True, doctype="<!DOCTYPE html>")
    OFFICIAL_WEBSITE = "An official website of the United States government"
    HOW_YOU_KNOW = "Here's how you know"
    BOARD_MANAGERS = "Board Manager Menu Users"
    CIAT_OCCM = "CIAT/OCCM Staff Menu Users"
    DEV_SA = "Developer/SysAdmin Menu Users"
    MENUS = (
        (BOARD_MANAGERS, "Board Managers"),
        (CIAT_OCCM, "CIAT/OCC"),
        (DEV_SA, "Developers"),
    )
    MENU_PARMS = {
        "glossaryconceptbydefinitionstatus.py": ["report"],
        "help.py": ["id"],
        "interventionandprocedureterms.py": ["IncludeAlternateNames"],
        "qcreport.py": ["DocType", "ReportType"],
        "summarymailerreport.py": ["flavor"],
    }

    def __init__(self, title, **kwargs):
        """Capture the initial settings for the page.

        Required positional argument:

            title
                string representing the page's browser title

        Optional keyword arguments:

            action
                form submission handler (default: URL which draws form)

            body_id
                ID attribute for body element (default: "cdr-page")

            head_title
                title for head block (default: page title)

            icon
                favicon path (defaults: "/favicon.ico")

            method
                set to "get" to override default of "post")

            scripts
                urls for js to load (default: jQuery and jQueryUI URLs)

            session
                object representing the current CDR login context

            subtitle
                used as the title of the page (not well named)

            stylesheets
                urls for CSS to load (default: CDR and jQueryUI CSS)
        """

        self.__title = title
        self.__opts = kwargs

    # ----------------------------------------------------------------
    # Instance methods.
    # ----------------------------------------------------------------

    def add_alert(self, message, **kwargs):
        """Add a prominent notification at the top of the main content.

        Required position argument:
          message - string or DOM object for base message; or sequence of same

        Optional keyword arguments:
          heading - optional string or DOM object for alert heading title
          type - string (info, warning, error, success; default: info)
          no_icon - if True, suppress display of alert icon
          slim - if True, decrease the height of the alert
          extra - optional sequence of additional elements to append to alert
        """

        container = self.main.find("div")
        if container is None or container.get("class") != "grid-container":
            raise Exception("add_alert(): grid container missing")
        alerts = container.find("div")
        if alerts is None or alerts.get("id") != "alerts-block":
            alerts = self.B.DIV(id="alerts-block")
            alerts.set("class", "padding-bottom-4")
            container.insert(0, alerts)
            self.main.set("class", "usa-section padding-top-4")
        alert_type = kwargs.get("type", "info")
        classes = ["usa-alert", f"usa-alert--{alert_type}"]
        if kwargs.get("no_icon"):
            classes.append("usa-alert--no-icon")
        if kwargs.get("slim"):
            classes.append("usa-alert--slim")
        body = self.B.DIV(self.B.CLASS("usa-alert__body"))
        alert = self.B.DIV(body, self.B.CLASS(" ".join(classes)))
        heading = kwargs.get("heading")
        if heading:
            body.append(self.B.H4(heading, self.B.STYLE("usa-alert__heading")))
        if message is not None:
            if not isinstance(message, (list, tuple)):
                message = [message]
            body.append(self.B.P(*message, self.B.CLASS("usa-alert__text")))
        extra = kwargs.get("extra", [])
        for item in extra:
            body.append(item)
        alerts.append(alert)

    def add_css(self, css):
        """Add style rules directly to the page."""
        self.head.append(self.B.STYLE(css))

    def add_output_options(self, default=None, onclick=None):
        """
        Allow the user to decide between HTML and Excel.
        """
        choices = ("html", "Web Page"), ("excel", "Excel Workbook")
        fieldset = self.fieldset("Report Format")
        fieldset.set("id", "report-format-block")
        self.form.append(fieldset)
        for value, label in choices:
            if value is None:
                value = ""
            opts = dict(label=label, value=value, onclick=onclick)
            if value == default:
                opts["checked"] = True
            fieldset.append(self.radio_button("format", **opts))
        return fieldset

    def add_uswds_script(self):
        """Separated out to be done at the last minute."""
        self.body.append(self.B.SCRIPT(src="/uswds/js/uswds.min.js"))

    def add_script(self, script):
        """Add script code directly to the page."""
        self.head.append(self.B.SCRIPT(script))

    def add_session_field(self, session):
        """Add hidden session field if it isn't there already."""

        if self.form is not None:
            if not self.form.xpath(f"//input[@name='{Controller.SESSION}']"):
                field = self.hidden_field(Controller.SESSION, session)
                self.form.append(field)

    def create_sidenav_item(self, positions, menu, labels):
        """Create a list item for the sidebar menu (possibly recursively).

        Required positional arguments:
          positions - sequence of integers for the nested menu positions
          menu - dictionary of values for the current menu item
          labels - sequence of labels for the menu path to the current page
                   (empty if that path can't be determined)

        Return:
          list item DOM object
        """

        item = self.B.LI(self.B.CLASS("usa-sidenav__item"))
        label = display = menu["label"]
        children = menu.get("children")
        current = False
        if labels and len(positions) <= len(labels):
            current = label == labels[len(positions)-1]
        if children and not current:
            display += " ..."
        link = self.B.A(display, href=self.find_menu_link(menu, positions))
        item.append(link)
        if current:
            link.set("class", "usa-current")
            if children:
                ul = self.B.UL(self.B.CLASS("usa-sidenav__sublist"))
                for i, child in enumerate(children):
                    args = positions + [i], child, labels
                    ul.append(self.create_sidenav_item(*args))
                item.append(ul)
        return item

    def filter_menu(self, menu):
        """Remove menu items the user is not allowed to use.

        Required positional argument:
          menu - nested dictionary of menu items

        Return:
          pruned menu or None if no actions are allowed
        """

        label = menu["label"]
        children = []
        for child in menu["children"]:
            if child.get("children"):
                child = self.filter_menu(child)
                if child:
                    children.append(child)
            else:
                permission = child.get("permission")
                if not permission or permission in self.user_permissions:
                    children.append(child)
                elif self.control:
                    args = child["label"], permission
                    self.control.logger.info("skipping %s (%s)", *args)
        if not children:
            return None
        return dict(label=label, children=children)

    def find_matching_path(self, menu):
        """Find menu path matching the current page.

        Required positional parameter:
            menu - dictionary of menu values

        Return:
            None or sequence of menu value dictionaries
        """

        script = menu.get("script")
        if script:
            if self.script_matches_page(script):
                return [menu]
            return None
        for child in menu.get("children", []):
            path = self.find_matching_path(child)
            if path:
                return [menu] + path
        return None

    def find_menu_link(self, menu, positions):
        script = menu.get("script")
        if script:
            separator = "&" if "?" in script else "?"
            key = "".join([f"{p:02x}" for p in positions])
            script = f"{script}{separator}_mp={key}"
            if self.control:
                script += f"&{Controller.SESSION}={self.control.session}"
            return script
        children = menu.get("children")
        if not children:
            return "javascript:void(0);"
        return self.find_menu_link(children[0], positions + [0])

    def menu_link(self, script, display, **params):
        """
        Add a list item containing a CDR admin menu link.
        """

        url = script
        if self.session and Controller.SESSION not in params:
            params[Controller.SESSION] = self.session
        if params:
            url = f"{url}?{urllib_parse.urlencode(params)}"
        return self.B.A(display, href=url)

    def script_matches_page(self, script):
        """Find out if a script matches the current page.

        Required positional argument:
          script - string for script name, possibly with parameters

        Return:
          True if the script matches the page, otherwise false
        """

        if not self.control or not self.control.script:
            return False
        if "?" in script:
            script, parms = script.split("?", 1)
        else:
            script, parms = script, ""
        self.control.logger.debug("page script is %s", self.control.script)
        script = script.lower()
        if self.control.script.lower() != script:
            return False
        if script not in self.MENU_PARMS:
            return True
        parms = urllib_parse.parse_qs(parms)
        for name in self.MENU_PARMS[script]:
            menu_value = parms.get(name)
            if menu_value and isinstance(menu_value, (list, tuple)):
                menu_value = menu_value[0]
            fields_value = self.control.fields.getvalue(name)
            if menu_value != fields_value:
                args = name, menu_value, fields_value
                self.control.logger.debug("parm %s: %s vs %s", *args)
                return False
        return True

    def send(self):
        """Push the page back to the browser via the web server."""

        if self.control:
            if self.control.show_news and self.news:
                for name in reversed(sorted(self.news, key=str.lower)):
                    news = self.news[name]
                    key = name.lower()
                    if "error" in key or "failure" in key:
                        message_type = "error"
                    elif "warning" in key:
                        message_type = "warning"
                    elif "success" in key:
                        message_type = "success"
                    else:
                        message_type = "info"
                    self.add_alert(news, type=message_type)
            if self.control.logged_out:
                opts = dict(href="/", style=f"color: {self.LINK_COLOR};")
                message = (
                    "You have been successfully logged out from the CDR.",
                    " Click ",
                    self.B.A("here", **opts),
                    " to log back in.",
                )
                self.add_alert(message, type="success")
        self.add_uswds_script()
        Controller.send_page(self.tostring())

    def tobytes(self):
        """Return the serialized page as ASCII bytes.

        The Unicode characters contained on the page will have
        been encoded as HTML entities.
        """

        return lxml_html.tostring(self.html, **self.STRING_OPTS)

    def tostring(self):
        """Return the serialized Unicode string for the page object."""

        opts = dict(self.STRING_OPTS, encoding="unicode")
        return lxml_html.tostring(self.html, **opts)

    # ----------------------------------------------------------------
    # Instance properties.
    # ----------------------------------------------------------------

    @cached_property
    def action(self):
        """URL for the form submission handler."""

        action = self.__opts.get("action")
        if action is None:
            action = Path(sys_argv[0]).name
        if action == "flask":
            action = ""
        return action

    @cached_property
    def banner(self):
        """USWDS standard banner."""

        img = self.B.IMG(src="/images/us_flag_small.png", alt="")
        img.set("class", "usa-banner__header-flag")
        img.set("aria-hidden", "true")
        flag = self.B.DIV(img, self.B.CLASS("grid-col-auto"))
        action = self.B.P(self.HOW_YOU_KNOW)
        action.set("class", "usa-banner__header-action")
        hidden = self.B.DIV(
            self.B.P(
                self.OFFICIAL_WEBSITE,
                self.B.CLASS("usa-banner__header-text")
            ),
            self.B.P(
                self.HOW_YOU_KNOW,
                self.B.CLASS("usa-banner__header-action")
            ),
            self.B.CLASS("grid-col-fill tablet:grid-col-auto")
        )
        hidden.set("aria-hidden", "true")
        span = self.B.SPAN(self.HOW_YOU_KNOW)
        span.set("class", "usa-banner__button-text")
        button = self.B.BUTTON(span, type="button")
        button.set("class", "usa-accordion__button usa-banner__button")
        button.set("aria-expanded", "false")
        button.set("aria-controls", "gov-banner-default")
        header = self.B.E(
            "header",
            self.B.DIV(
                flag,
                hidden,
                button,
                self.B.CLASS("usa-banner__inner")
            ),
            self.B.CLASS("usa-banner__header")
        )
        opts = dict(role="img", alt="")
        img = self.B.IMG(src="/images/icon-dot-gov.svg", **opts)
        img.set("class", "usa-banner__icon usa-media-block__img")
        img.set("aria-hidden", "true")
        dot_gov = self.B.DIV(
            img,
            self.B.DIV(
                self.B.P(
                    self.B.STRONG("Official websites use .gov"),
                    self.B.BR(),
                    "A ",
                    self.B.STRONG(".gov"),
                    " website belongs to an official government organization ",
                    "in the United States."
                ),
                self.B.CLASS("usa-media-block__body")
            ),
            self.B.CLASS("usa-banner__guidance tablet:grid-col-6")
        )
        desc = "Locked padlock icon"
        path_data = (
            "M26 0c10.493 0 19 8.507 19 19v9h3a4 4 0 0 1 4 4v28a4 4 0 0 1-4 "
            "4H4a4 4 0 0 1-4-4V32a4 4 0 0 1 4-4h3v-9C7 8.507 15.507 0 26 0zm0 "
            "8c-5.979 0-10.843 4.77-10.996 10.712L15 19v9h22v-9c0-6.075-4.925-"
            "11-11-11z"
        )
        path = self.B.E("path", fill="#000000", d=path_data)
        path.set("fill-rule", "evenodd")
        svg = self.B.E(
            "svg",
            self.B.TITLE("Lock", id="banner-lock-title"),
            self.B.E("desc", desc, id="banner-lock-description"),
            path,
            xmlns="http://www.w3.org/2000/svg",
            width="52",
            height="64",
            viewBox="0 0 52 64",
            role="img",
            focusable="false"
        )
        img = self.B.IMG(src="/images/icon-https.svg", **opts)
        img.set("class", "usa-banner__icon usa-media-block__img")
        img.set("aria-hidden", "true")
        svg.set("class", "usa-banner__lock-image")
        svg.set("aria-labelledby", "banner-lock-description")
        lock = self.B.DIV(
            img,
            self.B.DIV(
                self.B.P(
                    self.B.STRONG("Secure .gov websites use HTTPS"),
                    self.B.BR(),
                    "A ",
                    self.B.STRONG("lock"),
                    " ( ",
                    self.B.SPAN(svg, self.B.CLASS("icon-lock")),
                    " ) or ",
                    self.B.STRONG("https://"),
                    " means you’ve safely connected to the .gov website. ",
                    "Share sensitive information only on official, secure ",
                    "websites."
                ),
                self.B.CLASS("usa-media-block__body")
            ),
            self.B.CLASS("usa-banner__guidance tablet:grid-col-6")
        )
        content = self.B.DIV(
            self.B.DIV(dot_gov, lock, self.B.CLASS("grid-row grid-gap-lg")),
            self.B.CLASS("usa-banner__content usa-accordion__content")
        )
        content.set("id", "gov-banner-default")
        accordion = self.B.DIV(header, content, self.B.CLASS("usa-accordion"))
        banner = self.B.E("section", accordion, self.B.CLASS("usa-banner"))
        official_website = f"O{self.OFFICIAL_WEBSITE[4:]}"
        banner.set("aria-label", official_website)
        return banner

    @cached_property
    def body(self):
        """The body content element for the page."""

        body = self.B.BODY(id=self.body_id)
        if self.body_classes:
            body.set("class", " ".join(self.body_classes))
        skipnav = self.B.A("Skip to main content", href="#main-content")
        skipnav.set("class", "usa-skipnav")
        body.append(skipnav)
        body.append(self.banner)
        body.append(self.B.DIV(self.B.CLASS("usa-overlay")))
        body.append(self.header)
        body.append(self.main)
        body.append(self.footer)
        return body

    @cached_property
    def body_classes(self):

        classes = self.__opts.get("body_classes")
        if classes:
            if isinstance(classes, str):
                classes = classes.strip().split()
            return set(classes)
        return set()

    @cached_property
    def body_id(self):
        """ID attribute to be applied to the page's body element."""
        return self.__opts.get("body_id") or "cdr-page"

    @cached_property
    def control(self):
        """Control object for the page, if known."""
        return self.__opts.get("control")

    @cached_property
    def current_path(self):
        """Menu path to this page, if known."""

        if not self.control:
            return False
        if self.current_path_from_key:
            return self.current_path_from_key
        for menu in self.user_menus:
            path = self.find_matching_path(menu)
            if path:
                return path
        return None

    @cached_property
    def current_path_from_key(self):
        """Menu path derived from a parameter key."""

        key = self.control.fields.getvalue("_mp")
        if self.control:
            self.control.logger.info("current path key is %s", key)
        if not key or len(key) % 2 != 0:
            return None
        if not all(c in hexdigits for c in key):
            return None
        path = []
        menus = self.menus
        i = 0
        while i < len(key):
            if not menus:
                return None
            position = int(key[i:i+2], 16)
            if self.control:
                self.control.logger.debug("path position is %d", position)
            if position >= len(menus):
                return None
            menu = menus[position]
            path.append(menu)
            menus = menu.get("children")
            i += 2
        if not path:
            return None
        script = path[-1].get("script")
        if not self.script_matches_page(script):
            if self.control:
                self.control.logger.info("script %s not matched", script)
            return None
        return path

    @cached_property
    def cursor(self):
        """Read-only database cursor."""
        return db.connect(user="CdrGuest").cursor()

    @cached_property
    def footer(self):
        """Links at the bottom of the page."""
        return self.make_footer(self.session, self.user)

    @cached_property
    def form(self):
        """The body's <form> element."""

        opts = dict(method=self.method, action=self.action)
        form = self.B.FORM(self.B.CLASS("usa-form"), **opts)
        if self.session:
            form.append(self.hidden_field(Controller.SESSION, self.session))
        form.set("id", self.PRIMARY_FORM_ID)
        if self.__opts.get("body_classes") != "report":
            form.set("target", "_blank")
        return form

    @cached_property
    def head(self):
        """Assemble the head block for the HTML page."""

        http_equiv = self.B.META(content="IE=edge")
        http_equiv.set("http-equiv", "X-UA-Compatible")
        viewport = "width=device-width, initial-scale=1.0"
        head = self.B.HEAD(
            self.B.META(charset="utf-8"),
            http_equiv,
            self.B.META(name="viewport", content=viewport),
            self.B.META(
                name="description",
                content="CDR Administration Tools"
            ),
            self.B.META(
                name="author",
                content="Bob Kline and Volker Englisch"
            ),
            self.B.TITLE(self.head_title),
            self.B.LINK(href="/favicon.ico", rel="icon"),
            self.B.SCRIPT(src="/uswds/js/uswds-init.min.js")
        )
        for attrs in self.stylesheets:
            element = self.B.LINK()
            for key, value in attrs.items():
                element.set(key, value)
                if "rel" not in attrs:
                    element.set("rel", "stylesheet")
            head.append(element)
        for attrs in self.scripts:
            element = self.B.SCRIPT()
            for key, value in attrs.items():
                element.set(key, value)
            head.append(element)
        style = (
            "@media (min-width:30em){ .usa-form { max-width: none; }}",
            ".usa-nav__primary { margin-right: 1rem; }",
            ".usa-fieldset { margin-bottom: 2rem; }",
            ".usa-label { font-weight: bold; max-width: none; }",
            ".usa-form .usa-input--xl,",
            ".usa-form .usa-input-group--xl { max-width: none; }",
            ".usa-legend { font-size: 1.3em; font-weight: bold; }",
            ".usa-legend { max-width: none; }",
            ".usa-file-input { max-width: none; }",
            ".date-range-fields > div { display: inline-block; }",
            "body.report h1 { margin-bottom: 3rem; }",
            "table.usa-table { margin-bottom: 3rem; }",
            ".report-footer {",
            "  font-style: italic; font-size: .9em; text-align: center;",
            "}",
            ".report-footer #elapsed { color: green; }",
            ".hidden { display: none; }",
            ".nowrap { white-space: nowrap; }",
            ".usa-table td.text-bold { font-weight: bold; }",
            ".report .usa-table { width: 100%; }",
            ".report .usa-table td { word-wrap: break-word; }",
            ".break-all, .break-all * { word-break: break-all; }",
            f".usa-form a:visited {{ color: {self.LINK_COLOR}; }}",
            ".error { color: red; font-weight: bold; }",
        )
        head.append(self.B.STYLE("\n".join(style)))
        return head

    @cached_property
    def header(self):
        """The <header> element at the top of the body."""

        classes = "usa-nav__primary usa-accordion"
        primary_menu = self.B.UL(self.B.CLASS(classes))
        if len(self.user_menus) > 1:
            current_label = ""
            if self.current_path:
                current_label = self.current_path[0]["label"]
            labels = [m["label"] for m in self.menus]
            for user_menu in self.user_menus:
                link_class = "usa-nav-link"
                label = user_menu["label"]
                if label == current_label:
                    link_class += " usa-current"
                positions = [labels.index(label)]
                item = self.B.LI(
                    self.B.A(
                        self.B.SPAN(label),
                        self.B.CLASS(link_class),
                        href=self.find_menu_link(user_menu, positions)
                    ),
                    self.B.CLASS("usa-nav__primary-item")
                )
                primary_menu.append(item)
        search = self.B.E(
            "section",
            self.B.FORM(
                self.B.LABEL(
                    "Search",
                    self.B.FOR("search-field"),
                    self.B.CLASS("usa-sr-only")
                ),
                self.B.INPUT(
                    self.B.CLASS("usa-input"),
                    id="search-field",
                    type="search",
                    name="doc-id",
                    placeholder="CDR ID"
                ),
                self.B.INPUT(
                    type="hidden",
                    name=Controller.SESSION,
                    value=self.session.name
                ),
                self.B.BUTTON(
                    self.B.IMG(
                        self.B.CLASS("usa-search__submit-icon"),
                        src="/images/search--white.svg",
                        alt="Search"
                    ),
                    self.B.CLASS("usa-button"),
                    type="submit",
                ),
                self.B.CLASS("usa-search usa-search--small"),
                role="search",
                action="ShowCdrDocument.py",
                target="_blank"
            )
        )
        search.set("aria-label", "Search component")
        if not self.session or self.session.name == "guest":
            search.set("class", "margin-bottom-1")
        nav = self.B.E(
            "nav",
            self.B.BUTTON(
                self.B.IMG(src="/images/close.svg", role="img", alt="Close"),
                self.B.CLASS("usa-nav__close"),
                type="button"
            ),
            primary_menu,
            search,
            self.B.CLASS("usa-nav")
        )
        nav.set("aria-label", "Primary navigation")

        home = f"{self.BASE}/Admin.py?{Controller.SESSION}={self.session}"
        title = "CDR Administration"
        return self.B.E(
            "header",
            self.B.DIV(
                self.B.DIV(
                    self.B.DIV(
                        self.B.EM(self.B.A(title, title="Home", href=home)),
                        self.B.BUTTON(
                            "Menu",
                            self.B.CLASS("usa-menu-btn"),
                            type="button"
                        ),
                        self.B.CLASS("usa-logo"),
                        id="header-logo"
                    ),
                    self.B.CLASS("usa-navbar")
                ),
                nav,
                self.B.CLASS("usa-nav-container")
            ),
            self.B.CLASS("usa-header usa-header--basic")
        )

    @cached_property
    def head_title(self):
        """The title to be inserted into the head block of the page."""

        head_title = self.__opts.get("head_title")
        return self.title if head_title is None else head_title

    @cached_property
    def html(self):
        """Top-level object for the page."""
        return self.B.HTML(self.head, self.body)

    @cached_property
    def main(self):
        """This is where page content gets added."""

        container = self.B.DIV(self.B.CLASS("grid-container"))
        if self.__opts.get("body_classes") == "report" or self.sidenav is None:
            container.append(self.B.H1(self.subtitle))
            container.append(self.form)
        else:
            container.append(
                self.B.DIV(
                    self.B.DIV(
                        self.sidenav,
                        self.B.CLASS("tablet:grid-col-4")
                    ),
                    self.B.DIV(
                        self.B.H1(self.subtitle),
                        self.form,
                        self.B.CLASS("tablet:grid-col-8"),
                        id="report-form"
                    ),
                    self.B.CLASS("grid-row grid-gap")
                ),
            )
        return self.B.E("main", container, self.B.CLASS("usa-section"))

    @cached_property
    def menus(self):
        """Load the CDR administrative menu structures."""
        return self.load_menus(self.cursor)

    @cached_property
    def method(self):
        """CGI verb to be used for form submission."""
        return self.__opts.get("method", "post")

    @cached_property
    def news(self):
        """Information to be displayed at the top of the menu pages."""
        return getControlGroup("news")

    @cached_property
    def scripts(self):
        """Client-side scripts to be loaded for the page.

        To suppress the default scripts, pass an empty list to the constructor
        for the named 'scripts' argument.
        """

        scripts = self.__opts.get("scripts")
        if isinstance(scripts, (list, tuple)):
            script_attrs = []
            for script in scripts:
                if isinstance(script, str):
                    script_attrs.append(dict("src", script))
                elif isinstance(script, dict):
                    script_attrs.append(script)
                else:
                    self.session.logger.warning("bogus script %r", script)
            return script_attrs
        if scripts is not None:
            self.session.logger.warning("bogus scripts %r", scripts)
        #query = db.Query("ctl", "val")
        #query.where("grp = 'cdn'")
        #query.where("name = 'cgi-js'")
        #query.where("inactivated IS NULL")
        #row = query.execute(self.cursor).fetchone()
        #if row:
        #    return load_json_string(row.val)
        return self.SCRIPT_LINKS

    @cached_property
    def session(self):
        """CDR login context for this page."""

        session = self.__opts.get("session", "guest")
        return Session(session) if isinstance(session, str) else session

    @cached_property
    def sidenav(self):
        """Menu on the left sidebar."""

        if self.suppress_sidenav:
            return None
        ul = self.B.UL(self.B.CLASS("usa-sidenav"))
        menu_labels = [m["label"] for m in self.menus]
        if self.current_path:
            path_labels = [p["label"] for p in self.current_path]
            menu = self.current_path[0]
        else:
            return None
        if self.session:
            self.session.logger.info("path labels: %r", path_labels)
        positions = [menu_labels.index(menu["label"])]
        for i, child in enumerate(menu["children"]):
            args = positions + [i], child, path_labels
            ul.append(self.create_sidenav_item(*args))
        return self.B.E("nav", ul)

    @cached_property
    def subtitle(self):
        """Shown at the top of the page in an H1 element (badly named)."""
        return self.__opts.get("subtitle")

    @cached_property
    def suppress_sidenav(self):
        """Override to provide more nuanced logic."""
        return True if self.__opts.get("suppress_sidenav") else False

    @cached_property
    def stylesheets(self):
        """CSS rules to be loaded for the page."""
        return self.CSS_LINKS

    @cached_property
    def title(self):
        """Browser title for the page."""
        return self.__title or ""

    @cached_property
    def user(self):
        """Currently logged-in user."""
        return self.session.User(self.session, id=self.session.user_id)

    @cached_property
    def user_menus(self):
        """List of menus to which the current user has access."""

        menus = dict((m["label"], m) for m in self.menus)
        if not self.session:
            return [menus["Guest"]]
        allowed = []
        for group, label in self.MENUS:
            if group in self.user.groups:
                menu = self.filter_menu(menus[label])
                if menu:
                    allowed.append(menu)
        return allowed or [menus["Guest"]]

    @cached_property
    def user_permissions(self):
        """Set of actions the user is allowed to perform."""

        query = db.Query("grp_action ga", "a.name", "t.name")
        query.join("grp_usr gu", "gu.grp = ga.grp")
        query.join("action a", "a.id = ga.action")
        query.join("doc_type t", "t.id = ga.doc_type")
        query.where(query.Condition("gu.usr", self.user.id))
        rows = query.execute(self.session.cursor).fetchall()
        actions = set()
        for action, doctype in rows:
            if doctype:
                actions.add(f"{action}:{doctype}")
            else:
                actions.add(action)
        return actions


    # ----------------------------------------------------------------
    # Method not requiring an instance.
    # ----------------------------------------------------------------

    @staticmethod
    def load_menus(cursor=None):
        """Separated out so other tools can use it."""

        query = db.Query("ctl", "val")
        query.where("grp = 'admin'")
        query.where("name = 'menus'")
        query.where("inactivated IS NULL")
        row = query.execute(cursor).fetchone()
        if row:
            return load_json_string(row.val)
        directory = Path(__file__).parent
        path = directory / "menus.json"
        with path.open(encoding="utf-8") as fp:
            return load_json_file(fp)

    @classmethod
    def make_footer(cls, session=None, user=None):
        """Pull this out so it can be used from elsewhere.

        Optional keyword argument:
          session - current session for this request
          user - account which is currently logged in
        """

        li_classes = " ".join([
            "mobile-lg:grid-col-6",
            "desktop:grid-col-auto",
            "usa-footer__primary-content",
        ])
        link_class = "usa-footer__primary-link"
        session_name = session and str(session) or "guest"
        session_parm = f"{Controller.SESSION}={session_name}"
        link_values = [
            ("Help", f"HelpSearch.py?{session_parm}", False),
            ("NCI Web Site", "https://www.cancer.gov", True),
            ("CMS", "https://www-cms.cancer.gov", True),
            ("Filter", f"Filter.py?{session_parm}", False),
            ("Menus", f"show-menu-hierarchy.py?{session_parm}", True),
            ("Queries", f"CdrQueries.py?{session_parm}", True),
            ("Search", f"AdvancedSearch.py?{session_parm}", False),
        ]
        if session_name == "guest":
            values = "Log In", "/"
        else:
            values = "Log Out", f"LogOut.py?{session_parm}"
        link_values.append((*values, False))
        links = cls.B.UL(cls.B.CLASS("grid-row grid-gap"))
        for label, href, new_tab in link_values:
            link = cls.B.A(label, cls.B.CLASS(link_class), href=href)
            if new_tab:
                link.set("target", "_blank")
            if label == "Log Out":
                fullname = user.fullname
                name = user.name
                link.set("title", f"Logged in as {fullname} ({name}).")
            links.append(cls.B.LI(link, cls.B.CLASS(li_classes)))
        nav = cls.B.E(
            "nav",
            links,
            cls.B.CLASS("usa-footer__nav")
        )
        nav.set("aria-label", "footer navigation")
        return cls.B.E(
            "footer",
            cls.B.DIV(
                cls.B.DIV(
                    cls.B.DIV(
                        nav,
                        cls.B.CLASS("mobile-lg:grid-col-12")
                    ),
                    cls.B.CLASS("usa-footer__primary-container grid-row")
                ),
                cls.B.CLASS("usa-footer__primary-section")
            ),
            cls.B.CLASS("usa-footer usa-footer--slim")
        )


class Reporter:
    """Create web-based or Excel workbook reports.

    This replaces the older `Report` class.

    Uses the HTMLPage class for HTML report output, avoiding the ugly
    and error-prone approach of creating web pages by direct string
    juggling. When we have converted all of the existing reports to
    use this new class, move the inherited functions in here, rename
    `Reporter` to `Report` (here and for the users of this new class),
    and delete the old class.

    Example usage:

        R = cdrcgi.Reporter
        columns = (
            R.Column("Type ID", width="75px"),
            R.Column("Type Name", width="300px"),
        )
        query = db.Query("doc_type", "id", "name").order("name")
        rows = query.execute().fetchall()
        table = R.Table(rows, columns=columns, caption="Document Types")
        report = R("Simple Report", table)
        report.send("html")

    Look at ReportTemplate.py in the cgi directory for a more comprehensive
    guide to the features of this class and its nested classes.
    """

    def __init__(self, title, tables, **opts):
        """Capture the values for the report."""

        self.__title = title
        self.__tables = tables
        self.__opts = opts

    # ----------------------------------------------------------------
    # Instance method.
    # ----------------------------------------------------------------

    def send(self, format="html"):
        """Send the web page or Excel workbook to the browser.

        Optional positional argument:
          format - one of "html" (the default) or "excel"

        Return:
          Never returns, but exits (or perhaps throws an Exception)
        """

        if format == "excel":
            self.workbook.send()
            if self.debug:
                self.workbook.save(TMP)
        else:
            self.page.send()
        sys_exit(0)

    # ----------------------------------------------------------------
    # Instance properties.
    # ----------------------------------------------------------------

    @cached_property
    def css(self):
        """Sequence of string for <style> elements on HTML reports."""
        if not hasattr(self, "_css"):
            self._css = self.__opts.get("css")
            if isinstance(self._css, str):
                self._css = [self._css]
            elif not self._css:
                self._css = []
        return self._css

    @cached_property
    def debug(self):
        """Boolean; if True, save the workbook to the file system."""
        return self.__opts.get("debug")

    @cached_property
    def elapsed(self):
        """Optional length of time the report took to generate.

        Will be converted to a string at render time.
        """

        return self.__opts.get("elapsed")

    @cached_property
    def footer(self):
        """Optional footer to appear on the HTML report."""
        return self.__opts.get("footer")

    @cached_property
    def no_results(self):
        """What to say if no result tables are found (unused for Excel)."""

        if "no_results" in self.__opts:
            return self.__opts.get("no_results")
        return "No report results found."

    @cached_property
    def page(self):
        """HTML version of report."""

        opts = {
            "subtitle": self.subtitle,
            "body_classes": "report",
            "styles": self.css,
        }
        opts.update(self.page_opts)
        page = HTMLPage(self.title, **opts)
        if self.css:
            page.add_css(self.css)
        if not self.tables and self.no_results:
            no_results = page.B.P(self.no_results)
            no_results.set("class", "no-results")
            page.form.append(no_results)
        for table in self.tables:
            if table.node is not None:
                page.form.append(table.node)
        if self.footer is not None:
            page.main.append(self.footer)
        if self.elapsed:
            footnote = page.B.P(f"elapsed: {self.elapsed}")
            footnote.set("class", "footnote")
            page.main.append(footnote)
        return page

    @cached_property
    def page_opts(self):
        """Dictionary of options passed to HTMLPage class's constructor."""
        return self.__opts.get("page_opts") or {}

    @cached_property
    def subtitle(self):
        """Title shown at the top of the page (not well named).

        Not used for Excel reports.
        """

        return self.__opts.get("subtitle")

    @cached_property
    def tables(self):
        """Sequence of tables to be included in the report.

        A single table will be converted to a list.
        """

        if isinstance(self.__tables, Reporter.Table):
            return [self.__tables]
        return self.__tables

    @cached_property
    def title(self):
        """Title of the report.

        Used for the workbook filename or the /html/head/title element.
        """

        return self.__title

    @cached_property
    def workbook(self):
        """Wrapper for Excel workbook."""

        workbook = Excel(self.title, stamp=True, wrap=self.wrap)
        for table in self.tables:
            table.add_worksheet(workbook)
        return workbook

    @cached_property
    def wrap(self):
        return self.__opts.get("wrap", True)

    class Cell:
        """Data for one cell in a report table."""

        B = html_builder

        def __init__(self, *values, **opts):
            """Capture the information needed to show this cell on the report.
            """

            self.__vals = values
            self.__opts = opts

        def write(self, book, rownum, colnum, columns):
            """Add this cell's data to the worksheet.

            Only used for the Excel flavor of the report.

            Pass:
                book - `Excel` object
                rownum - integer position for the row we're on
                colnum - starting column for this cell's data
                columns - sequence of `Reporter.Column` objects for the table

            Return:
                integer position for the next cell to be written
            """

            # May have to skip columns if there were previous rowspans.
            while colnum <= len(columns):
                if columns[colnum-1].skip > 0:
                    columns[colnum-1].skip -= 1
                    colnum += 1
                else:
                    break
            if colnum > len(columns):
                raise Exception(f"too many cells for row {rownum:d}")

            # Assemble the values to be written
            if len(self.values) == 1:
                values = self.values[0]
            else:
                values = "\n".join([str(value) for value in self.values])

            # Determine the styles to be applied.
            styles = self.sheet_styles
            if self.right:
                styles["alignment"] = book.right
            elif self.center:
                styles["alignment"] = book.center
            elif "alignment" not in styles:
                styles["alignment"] = book.left
            if self.href:
                values = f'=HYPERLINK("{self.href}", "{values}")'
                styles["font"] = book.hyperlink
            elif self.bold:
                styles["font"] = book.bold

            # Handle the case of writing to more than one cell.
            nextcol = colnum + 1
            if self.colspan or self.rowspan:
                colspan = self.colspan or 1
                rowspan = self.rowspan or 1
                r1, c1 = r2, c2 = rownum, colnum
                if colspan > 1:
                    c2 += colspan - 1
                    if c2 > len(columns):
                        msg = "not enough room for colspan on row {rownum:d}"
                        raise Exception(msg)
                if rowspan > 1:
                    extra_rows = rowspan - 1
                    r2 += extra_rows
                    while colnum <= c2:
                        if columns[colnum-1].skip > 0:
                            msg = "overlapping rowspan at row {rownum:d}"
                            raise Exception(msg)
                        columns[colnum-1].skip = extra_rows
                        colnum += 1
                nextcol = c2 + 1
                book.merge(r1, c1, r2, c2)
                colnum = c1
            # This resulted in horrible performance. Use formula instead.
            # cell = book.write(rownum, colnum, values, styles)
            # if self.href:
            #     cell.hyperlink = self.href
            if isinstance(values, bytes):
                try:
                    values = str(values, encoding="utf-8")
                except Exception:
                    values = str(values)
            book.write(rownum, colnum, values, styles)
            return nextcol

        @cached_property
        def bold(self):
            """Boolean indicating whether the cell values should be bolded."""
            return self.__opts.get("bold")

        @cached_property
        def center(self):
            """Boolean indicating whether to center the cell's content."""
            return self.__opts.get("center")

        @cached_property
        def classes(self):
            """Sequence of class names for this cell."""

            classes = self.__opts.get("classes")
            if not classes:
                classes = set()
            elif isinstance(classes, set):
                classes = classes
            elif isinstance(classes, str):
                classes = {classes}
            elif isinstance(classes, (tuple, list)):
                classes = set(classes)
            else:
                message = "unexpected type {} for Cell classes: {}"
                args = type(classes), repr(classes)
                raise Exception(message.format(args))
            if self.bold and not self.href:
                classes.add("text-bold")
            if self.center:
                classes.add("text-center")
            elif self.right:
                classes.add("text-right")
            if self.middle:
                classes.add("text-middle")
            return classes

        @cached_property
        def colspan(self):
            """How many columns does this cell span horizontally?"""

            colspan = self.__opts.get("colspan")
            return int(colspan) if colspan else None

        @cached_property
        def href(self):
            """URL value for link."""
            return self.__opts.get("href")

        @cached_property
        def middle(self):
            """Override vertical alignment."""
            return True if self.__opts.get("middle") else False

        @cached_property
        def right(self):
            """True if values should be right aligned."""
            return True if self.__opts.get("right") else False

        @cached_property
        def rowspan(self):
            """How many rows does this cell span vertically?"""

            rowspan = self.__opts.get("rowspan")
            return int(rowspan) if rowspan else None

        @cached_property
        def sheet_styles(self):
            """Optional dictionary of style attributes for the Excel report."""
            return self.__opts.get("sheet_styles") or {}

        @cached_property
        def style(self):
            """Custom CSS specified directly on the element.

            Not a best practice, but needed to work around bugs in
            Microsoft Word.
            """

            return self.__opts.get("style")

        @cached_property
        def target(self):
            """Target for links.

            Only used for web output, when the `href` option has been set.
            """

            return self.__opts.get("target")

        @cached_property
        def td(self):
            """HTML node for this cell in a web report."""

            td = container = self.B.TD()
            if self.href:
                container = self.B.A(href=self.href)
                if self.target:
                    container.set("target", self.target)
                if self.bold:
                    container.set("class", "strong")
                td.append(container)
            if len(self.values) == 1:
                if isinstance(self.values[0], HtmlElement):
                    container.append(self.values[0])
                else:
                    container.text = str(self.values[0])
            else:
                values = list(self.values)
                container.append(self.B.SPAN(str(values.pop(0))))
                while values:
                    container.append(self.B.BR())
                    container.append(self.B.SPAN(str(values.pop(0))))
            if self.colspan:
                td.set("colspan", str(self.colspan))
            if self.rowspan:
                td.set("rowspan", str(self.rowspan))
            if self.classes:
                td.set("class", " ".join(self.classes))
            if self.title:
                td.set("title", self.title)
            if self.style:
                td.set("style", self.style)
            return td

        @cached_property
        def title(self):
            """Optional string to be shown in popup when hovering."""

            if "title" in self.__opts:
                return self.__opts["title"]
            elif "tooltip" in self.__opts:
                return self.__opts["tooltip"]
            return None

        @cached_property
        def tooltip(self):
            """Alias for the title property."""
            return self.title

        @cached_property
        def values(self):
            """Collect the values for the cell into a sequence."""

            values = []
            for value in self.__vals:
                if isinstance(value, (list, tuple)):
                    values += list(value)
                elif value is not None:
                    values.append(value)
            return values or [""]

    class Column:
        """Header and properties for one column in a report table."""

        def __init__(self, name, **opts):
            """Save what we need to render this column in the report."""
            self.__name = name
            self.__opts = opts

        @cached_property
        def classes(self):
            """Optional classes for the th element (HTML only)."""
            return self.__opts.get("classes")

        @cached_property
        def colspan(self):
            """How many columns does this header need to cover?"""
            return self.__opts.get("colspan")

        @cached_property
        def id(self):
            """Optional unique ID for the th element (HTML only)."""
            return self.__opts.get("id")

        @cached_property
        def name(self):
            """What we display at the top of the column."""
            return self.__name

        @cached_property
        def skip(self):
            """Keep track of rows to be skipped when rowspan is set."""
            return 0

        @cached_property
        def tooltip(self):
            """Popup string for help when the user hovers over the column."""
            return self.__opts.get("tooltip")

        @cached_property
        def style(self):
            """HTML style attribute for column element on web page."""

            style = self.__opts.get("style") or ""
            rules = [r for r in style.rstrip(";").split(";") if r]
            if self.width:
                rules.append(f"min-width: {self.width}")
            return ";".join(rules) if rules else None

        @cached_property
        def width(self):
            """Minimum width of column (e.g., '40px')."""
            return self.__opts.get("width")

    class Table:
        """Grid of rows and columns for the report.

        A report can have more than one table.
        """

        B = html_builder
        WIDE_CSS = (
            ".report .usa-table { width: 90%; margin: 3rem auto 1.25rem; }"
        )

        def __init__(self, rows, **opts):
            """Capture the information we need to render the table."""
            self.__rows = rows
            self.__opts = opts

        def add_worksheet(self, book):
            """Create an Excel worksheet and add it to the workbook.

            Pass:
                book
                    `Excel` object
            """

            # Create a new worksheet.
            self.sheet = book.add_sheet(self.sheet_name)
            self.sheet.sheet_format.defaultRowHeight = 200
            if self.freeze_panes:
                self.sheet.freeze_panes = self.freeze_panes

            # Add the rows for the caption strings.
            rownum = 1
            styles = dict(alignment=book.center, font=book.bold)
            for caption in self.caption:
                book.merge(rownum, 1, rownum, len(self.columns))
                book.write(rownum, 1, caption, styles)
                rownum += 1
            if self.caption:
                book.merge(rownum, 1, rownum, len(self.columns))
                rownum += 1

            # Show the report date between caption and headers if requested.
            if self.show_report_date:
                report_date = f"Report date: {date.today()}"
                book.merge(rownum, 1, rownum, len(self.columns))
                book.write(rownum, 1, report_date, dict(font=book.bold))
                rownum += 1
                book.merge(rownum, 1, rownum, len(self.columns))
                rownum += 1

            # Set the column headers and widths.
            colnum = 1
            styles["alignment"] = book.center_middle
            for column in self.columns:
                if column.width:
                    width = book.pixels_to_chars(column.width)
                    book.set_width(colnum, width)
                book.write(rownum, colnum, column.name, styles)
                colnum += 1

            # Add each of the data rows to the worksheet.
            for row in self.rows:
                rownum += 1
                colnum = 1
                for cell in row:
                    if not isinstance(cell, Reporter.Cell):
                        cell = Reporter.Cell(cell)
                    colnum = cell.write(book, rownum, colnum, self.columns)

        def debug(self, message, *args):
            """If we have a logger, use it for debugging."""

            if self.logger:
                self.logger.debug(message, *args)

        @cached_property
        def caption(self):
            """Sequence of strings to be displayed for the table's caption.

            If more than one string, each will be rendered on a separate line.
            """

            caption = self.__opts.get("caption")
            if not caption:
                return []
            return [caption] if isinstance(caption, str) else caption

        @cached_property
        def classes(self):
            """Optional classes for the table element (HTML only)."""
            return self.__opts.get("classes")

        @cached_property
        def cols(self):
            """Alias for `columns` property."""
            return self.columns

        @cached_property
        def columns(self):
            """Sequence of `Reporter.Column` objects for this table."""

            values = self.__opts.get("columns") or self.__opts.get("cols")
            columns = []
            if values:
                for value in values:
                    if isinstance(value, str):
                        value = Reporter.Column(value)
                    columns.append(value)
            return columns

        @cached_property
        def fixed(self):
            """True if the table layout should be fixed."""
            return True if self.__opts.get("fixed") else False

        @cached_property
        def freeze_panes(self):
            """Optional cell marking row/col freezing (Excel only)."""
            return self.__opts.get("freeze_panes")

        @cached_property
        def id(self):
            """Optional id attribute for the table element (HTML only)."""
            return self.__opts.get("id")

        @cached_property
        def logger(self):
            """Access to logging."""
            return self.__opts.get("logger")

        @cached_property
        def node(self):
            """HTML object for table."""

            # Hold off until we know we have child nodes for the table.
            children = []

            # Add the caption strings to the table if we have any.
            caption = list(self.caption)
            if self.show_report_date:
                caption.append("")
                caption.append(f"Report date: {date.today()}")
            if caption:
                nodes = [self.B.SPAN(caption[0])]
                for line in caption[1:]:
                    nodes.append(self.B.BR())
                    nodes.append(self.B.SPAN(line))
                children.append(self.B.CAPTION(*nodes))

            # Add the column headers if they have been provided.
            if self.columns:
                colgroup = self.B.COLGROUP()
                children.append(colgroup)
                tr = self.B.TR()
                for column in self.columns:
                    col = self.B.COL()
                    if column.width:
                        col.set("style", f"width: {column.width}")
                    colgroup.append(col)
                    th = self.B.TH(column.name)
                    if column.style:
                        th.set("style", column.style)
                    if column.id:
                        th.set("id", column.ids)
                    if column.tooltip:
                        th.set("title", column.tooltip)
                    if column.classes:
                        if isinstance(column.classes, str):
                            th.set("class", column.classes)
                        else:
                            th.set("class", " ".join(column.classes))
                    if column.colspan:
                        th.set("colspan", str(column.colspan))
                        col.set("span", str(column.colspan))
                    tr.append(th)
                children.append(self.B.THEAD(tr))

            # Only create the <tbody> element if there are data rows.
            if self.rows:
                tbody = self.B.TBODY()
                for row in self.rows:
                    tr = self.B.TR()
                    self.debug("row has %d cells", len(row))
                    for cell in row:
                        if not isinstance(cell, Reporter.Cell):
                            cell = Reporter.Cell(cell)
                        self.debug("cell values: %s", cell.values)
                        self.debug("td: %s", cell.td)
                        tr.append(cell.td)
                    self.debug("tr: %s", lxml_html.tostring(tr))
                    tbody.append(tr)
                children.append(tbody)

            # Create the table element if we found any child nodes.
            if not children:
                return None
            node = self.B.TABLE(*children)
            if self.fixed:
                node.set("style", "table-layout: fixed;")
            if self.id:
                node.set("id", self.id)
            classes = self.classes
            if not classes:
                classes = set()
            elif isinstance(classes, str):
                classes = {classes}
            classes.add("usa-table")
            classes.add("usa-table--borderless")
            node.set("class", " ".join(classes))
            return node

        @cached_property
        def rows(self):
            """Sequence of data cells displayed by the table."""
            return self.__rows

        @cached_property
        def sheet_name(self):
            """Optional name of the sheet.

            Not used for HTML reports. Defaults to "SheetN" where
            N is the number of existing sheets plus 1.
            """

            return self.__opts.get("sheet_name")

        @cached_property
        def show_report_date(self):
            """If True, add the report date between the caption and headers."""
            return True if self.__opts.get("show_report_date") else False


class BasicWebPage:
    """Avoids the USWDS framework to accomodate wider report tables."""

    B = html_builder
    CSS = (
        'body { font-family: "Source Sans Pro Web", Arial, sans-serif; }',
        "body > div { width: 95%; margin: 2rem auto; }",
        "table { border-collapse: collapse; }",
        "table caption { font-weight: bold; padding: 1rem; }",
        "th, td { font-size: .95em; border: 1px solid black; padding: .25rem}",
        "th { vertical-align: middle; }",
        "td { vertical-align: top; }",
        ".report-footer { font-style: italic; font-size: .9em; ",
        "                 text-align: center; }",
        "#elapsed { color: green; }",
        ".error { color: red; font-weight: bold; }",
        ".nowrap { white-space: nowrap; }",
        ".hidden { display: none; }",
        ".text-center { text-align: center; }",
     )

    def send(self):
        """Serialize the DOM and send it back to the browser."""
        Controller.send_page(self.page)

    @cached_property
    def body(self):
        """DOM object for the HTML BODY block."""
        return self.B.BODY(self.wrapper)

    @cached_property
    def head(self):
        """DOM object for the HTML HEAD block."""

        http_equiv = self.B.META(content="IE=edge")
        http_equiv.set("http-equiv", "X-UA-Compatible")
        viewport = "width=device-width, initial-scale=1.0"
        return self.B.HEAD(
            self.B.META(charset="utf-8"),
            http_equiv,
            self.B.META(name="viewport", content=viewport),
            self.B.TITLE("CDR Administration"),
            self.B.LINK(href="/favicon.ico", rel="icon"),
            self.B.STYLE("\n".join(self.CSS))
        )

    @cached_property
    def page(self):
        """DOM for the report's HTML page."""
        return self.B.HTML(self.head, self.body)

    @cached_property
    def wrapper(self):
        """DOM object for the HTML DIV element around body content."""
        return self.B.DIV()


class Excel:
    """Build workbooks using OpenPyXl

    See https://openpyxl.readthedocs.io/en/stable/

    We have created most of our Excel reports in the past using
    the pre-2007 formats. This package uses the current, more
    capable Excel format, and is the recommended package for
    generating Excel workbooks in Python.

    In rare cases it will be necessary to use the XlsxWriter
    package (https://xlsxwriter.readthedocs.io). See, for example,
    the MediaCaptionContent.py web admin report. One drawback of
    the XlsxWriter package is that it cannot read Excel workbooks,
    only create them.
    TODO: document the limitations of the OpenPyXl package which
    require the use of XlsxWriter.

    Note that an object of this class has properties for commonly
    used styles. It caches these properties, rather that creating
    a new style object for each call, so be careful that you not
    alter the returned property unless you are sure that you will
    want _all_ uses of that style property to have that change.
    """

    MIME_SUBTYPE = "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    MIME_TYPE = f"application/{MIME_SUBTYPE}"
    WINDOW_WIDTH = 25000
    WINDOW_HEIGHT = 15000
    from openpyxl.utils import get_column_letter

    def __init__(self, title=None, **opts):
        self.__title = title
        self.__opts = opts

    def add_sheet(self, name=None):
        if not name:
            name = f"Sheet{len(self.book.worksheets)+1}"
        sheet = self.book.create_sheet(title=name)
        sheet.print_options.gridLines = True
        return sheet

    def merge(self, start_row, start_column, end_row, end_column):
        self.sheet.merge_cells(
            start_row=start_row,
            start_column=start_column,
            end_row=end_row,
            end_column=end_column
        )

    def save(self, directory=None):
        path = self.filename
        if directory:
            path = f"{directory}/{path}"
        with open(path, "wb") as fp:
            self.book.save(fp)
        return path

    def send(self):
        headers = (
            f"Content-type: {self.book.mime_type}",
            f"Content-disposition: attachment; filename={self.filename}"
        )
        headers = "\r\n".join(headers) + "\r\n\r\n"
        sys_stdout.buffer.write(headers.encode("utf-8"))
        self.book.save(sys_stdout.buffer)

    def set_width(self, col, width):
        letter = Excel.get_column_letter(col)
        self.sheet.column_dimensions[letter].width = width

    def write(self, row, column, value, styles=None):
        cell = self.sheet.cell(row=row, column=column, value=value)
        if styles:
            for name in styles:
                setattr(cell, name, styles[name])
        return cell

    @cached_property
    def hyperlink(self):
        """Font styling we want for links."""

        opts = dict(color="000000FF", underline="single")
        return excel_styles.Font(**opts)

    @cached_property
    def bold(self):
        return excel_styles.Font(bold=True)

    @cached_property
    def center(self):

        opts = dict(horizontal="center", vertical="top", wrap_text=self.wrap)
        return excel_styles.Alignment(**opts)

    @cached_property
    def right(self):

        opts = dict(horizontal="right", vertical="top", wrap_text=self.wrap)
        return excel_styles.Alignment(**opts)

    @cached_property
    def left(self):

        opts = dict(horizontal="left", vertical="top", wrap_text=self.wrap)
        return excel_styles.Alignment(**opts)

    @cached_property
    def center_middle(self):
        """Center both horizontally and vertically."""

        opts = dict(
            horizontal="center",
            vertical="center",
            wrap_text=self.wrap,
        )
        self._center_middle = excel_styles.Alignment(**opts)

    @cached_property
    def book(self):
        """Create a workbook with no sheets."""

        book = ExcelWorkbook()
        for view in book.views:
            view.windowWidth = self.WINDOW_WIDTH
            view.windowHeight = self.WINDOW_HEIGHT
        if not self.__opts.get("keep_initial_sheet"):
            for sheet in book.worksheets:
                book.remove(sheet)
        return book

    @cached_property
    def sheet(self):
        """The most recently added sheet if any, else None."""
        return self.book.worksheets[-1] if self.book.worksheets else None

    @cached_property
    def title(self):
        return self.__title

    @cached_property
    def filename(self):

        stamp = ""
        if self.__opts.get("stamp") or not self.title:
            stamp = datetime.now().strftime("-%Y%m%d%H%M%S")
        if self.title:
            title = re_sub(r"\W", "_", self.title)
        else:
            title = "workbook"
        return f"{title}{stamp}.xlsx"

    @cached_property
    def wrap(self):
        return self.__opts.get("wrap")

    @classmethod
    def pixels_to_chars(cls, pixels):
        """Convert pixels to characters for column widths, etc.

        See https://docs.microsoft.com/en-us/previous-versions/office\
        /developer/office-2010/cc802410(v=office.14)

        In particular:
          "To translate from pixels to character width, use this calculation:
            =Truncate(({pixels}-5)/{Maximum Digit Width} * 100+0.5)/100

        The following code gets the maximum digit width for the default
        font in a new Excel installation:

        >>> from PIL import ImageFont
        >>> font = ImageFont.truetype('d:/Windows/Fonts/calibri.ttf', 11)
        >>> font.getsize("0")
        (6, 9)

        In other words, 6 pixels wide.

        To go back from chars to pixels:
            pixels = int(round(6 * ((100 * chars - .5) / 100) + 5))
        """

        max_digit_width = 6
        pixels = int(re_sub("[^0-9]+", "", pixels))
        return int((pixels - 5) / max_digit_width * 100 + 0.5) / 100


class AdvancedSearch(FormFieldFactory):
    """Search for CDR documents of a specific type.

    It is the responsibility of the derived classes to:
       1. populate the search_fields member
       2. populate the query_fields member
       3. assign the class-level DOCTYPE value
       4. assign the class-level FILTER value (if needed)
       5. override customize_form() if appropriate
       6. override customize_report() if appropriate

    Public methods:

        run()
            top-level processing entry point
    """

    INCLUDE_ROWS = True
    FILTER = SUBTITLE = DOCTYPE = None
    DBQuery = db.Query
    SESSION = Controller.SESSION
    REQUEST = Controller.REQUEST
    BASE = Controller.BASE

    def __init__(self):
        self.match_all = True if self.fields.getvalue("match_all") else False
        self.session = Session(self.fields.getvalue(self.SESSION) or "guest")
        self.request = self.fields.getvalue(self.REQUEST)
        self.search_fields = []
        self.query_fields = []

    def run(self):
        try:
            if self.request == "Search":
                self.show_report()
            else:
                self.show_form()
        except Exception as e:
            try:
                message = "AdvancedSearch.run(request=%r)"
                self.logger.exception(message, self.request)
            except Exception:
                Controller.bail(f"Unable to log exception {e}")
            Controller.bail(f"AdvancedSearch.run() failure: {e}")

    def show_form(self, subtitle=None, error=None):
        args = self.session.name, subtitle or self.SUBTITLE, self.search_fields
        page = self.Form(*args, error=error, control=self)
        self.customize_form(page)
        classes = page.B.CLASS("button usa-button")
        opts = dict(name="Request", value="Search", type="submit")
        submit = page.B.INPUT(classes, **opts)
        page.form.append(submit)
        page.send()

    def show_report(self):
        args = self.query_fields, self.DOCTYPE, self.match_all
        self.query = query = self.Query(*args)
        self.rows = rows = query.execute(self.session.cursor).fetchall()
        connector = " and " if self.match_all else " or "
        strings = connector.join([repr(c) for c in query.criteria])
        subtitle = self.SUBTITLE
        opts = dict(search_strings=strings, count=len(rows), subtitle=subtitle)
        opts["control"] = self
        opts["session"] = self.session
        if self.INCLUDE_ROWS:
            opts["rows"] = rows
        if self.FILTER:
            opts["filter"] = self.FILTER
        page = self.ResultsPage(self.DOCTYPE, **opts)
        self.customize_report(page)
        page.send()

    def customize_form(self, page):
        """Override in derived class if default behavior isn't enough."""

    def customize_report(self, page):
        """Override in derived class if default behavior isn't enough."""

    def values_for_paths(self, paths):
        query = db.Query("query_term", "value").unique().order("value")
        query.where(query.Condition("path", paths, "IN"))
        rows = query.execute(self.session.cursor).fetchall()
        return [row.value for row in rows if row.value.strip()]

    @cached_property
    def countries(self):
        query = db.Query("document d", "d.id", "d.title")
        query.join("doc_type t", "t.id = d.doc_type")
        query.where("t.name = 'Country'")
        query.order("d.title")
        rows = query.execute(self.session.cursor).fetchall()
        return [(f"CDR{row.id:010d}", row.title) for row in rows]

    @cached_property
    def fields(self):
        """Named values from the CGI form."""
        return FieldStorage()

    @cached_property
    def logged_out(self):
        """True if the user has just logged out."""
        return True if self.fields.getvalue("logged_out") else False

    @cached_property
    def logger(self):
        """How we record what we do."""
        return self.session.logger

    @cached_property
    def script(self):
        """Name of form submission handler."""
        return Path(sys_argv[0]).name

    @cached_property
    def show_news(self):
        """Whether we should display news announcements."""
        return True if self.fields.getvalue("show_news") else False

    @cached_property
    def states(self):
        fields = "s.id AS i", "s.title AS s", "c.title as c"
        query = db.Query("document s", *fields)
        query.join("query_term r", "r.doc_id = s.id")
        query.join("document c", "c.id = r.int_val")
        query.where("r.path = '/PoliticalSubUnit/Country/@cdr:ref'")
        query.order("s.title", "c.title")
        rows = query.execute(self.session.cursor).fetchall()
        return [(f"CDR{row.i:010d}", f"{row.s} [{row.c}]") for row in rows]

    @cached_property
    def statuses(self):
        """Valid active_status value for non-deleted CDR documents."""
        return [("A", "Active"), ("I", "Inactive")]

    @cached_property
    def valid_values(self):
        return dict(getDoctype("guest", self.DOCTYPE).vvLists)

    class Form(HTMLPage):
        TITLE = "CDR Advanced Search"
        MATCH_ALL_HELP = "If unchecked any field match will succeed."

        def __init__(self, session, subtitle, fields, **kwargs):
            kwargs = dict(
                **kwargs,
                session=session,
                subtitle=subtitle,
                method="get",
            )
            HTMLPage.__init__(self, self.TITLE, **kwargs)
            self.add_session_field(session or "guest")
            fieldset = self.fieldset("Search Fields")
            if kwargs.get("error"):
                classes = self.B.CLASS("error center")
                self.form.append(self.B.P(kwargs["error"], classes))
            self.form.append(fieldset)
            for field in fields:
                fieldset.append(field)
            if len(fields) > 1:
                fieldset = self.fieldset("Options")
                self.form.append(fieldset)
                opts = dict(
                    label="Match All Criteria",
                    checked=True,
                    value="yes",
                    tooltip=self.MATCH_ALL_HELP,
                )
                fieldset.append(self.checkbox("match_all", **opts))
            else:
                self.form.append(self.hidden_field("match_all", "yes"))

    class ResultsPage(HTMLPage):

        TITLE = "CDR Advanced Search Results"
        PERSONS_ORGS_GLOSSARY = (
            "Person",
            "Organization",
            "GlossaryTermConcept",
            "GlossaryTermName",
        )

        def __init__(self, doctype, **kwargs):

            # Let the base class get us started.
            opts = dict(
                body_id="advanced-search-results",
                subtitle=kwargs.get("subtitle"),
                session=kwargs.get("session"),
                control=kwargs.get("control"),
            )
            HTMLPage.__init__(self, self.TITLE, **opts)

            # Start the table for the results.
            table = self.B.TABLE()
            table.set("class", "usa-table usa-table--borderless")
            self.add_css("#advanced-search-results table { min-width: 100%; }")

            # Add a summary line if we have the requisite information.
            strings = kwargs.get("search_strings")
            count = kwargs.get("count")
            if strings and isinstance(count, int):
                if count == 1:
                    summary = f"1 document matches {strings}"
                else:
                    summary = f"{count:d} documents match {strings}"
                table.append(self.B.CAPTION(summary))

            # Attach the table to the page.
            self.form.append(table)

            # If we have the rows for the results set, add them to the table.
            rows = kwargs.get("rows")
            if rows:
                for i, row in enumerate(rows):
                    cdr_id = f"CDR{Doc.extract_id(row[0]):010d}"
                    title = row[1].replace(";", "; ")

                    # Create the link.
                    base = f"{Controller.BASE}/QcReport.py"
                    if doctype in self.PERSONS_ORGS_GLOSSARY:
                        url = f"{base}?DocId={cdr_id}"
                    elif doctype == "Summary":
                        url = f"{base}?DocId={cdr_id}&ReportType=nm"
                    else:
                        base = f"{Controller.BASE}/Filter.py"
                        filtre = kwargs.get("filter")
                        url = f"{base}?DocId={cdr_id}&Filter={filtre}"
                    url += f"&{Controller.SESSION}={self.session}"
                    link = self.B.A(cdr_id, href=url)

                    # Assemble the table row and attach it.
                    tr = self.B.TR(self.B.CLASS("row-item"))
                    tr.append(self.B.TD(f"{i+1}.", self.B.CLASS("text-right")))
                    tr.append(self.B.TD(link, self.B.CLASS("text-center")))
                    tr.append(self.B.TD(title, self.B.CLASS("doc-title")))
                    table.append(tr)

        @cached_property
        def sidenav(self):
            return None

    class QueryField:
        """Information used to plug one field into the search query."""

        def __init__(self, var, selectors):
            """Capture the value and the paths used to look for it.

            Pass:
                var - the CGI variable's value for this search field
                selectors - list of paths to look for in the query_term table

            In some cases `selectors` is a single string, in which case it
            is the name of a column in the `document` table.
            """

            self.var = var
            self.selectors = selectors

    class Query:
        """Builds a database query for an advanced search.

        Attributes:
            fields - information on the search fields from the form
            doctype - string identifying which documents we're looking for
            match_all - if False then OR the field conditions together
            criteria - list of all the field values being search for,
                       so we can remind the user what she put on the form
            query - db.Query object built on demand
        """

        def __init__(self, fields, doctype, match_all=True):
            """Capture what we'll need to build the SQL query.

            Pass:
                fields - sequence of `QueryField` objects
                doctype - string for the CDR document type (e.g., 'Summary')
                match_all - if set to `False` then _any_ field match works
            """

            self.fields = fields
            self.match_all = match_all
            self.doctype = doctype
            self.criteria = []

        def execute(self, cursor=None):
            """Convenience method to make the code cleaner."""
            return self.query.execute(cursor)

        @cached_property
        def query(self):
            """Build the query for the search if it's not already cached."""

            # Set up some convenience aliases.
            Query = db.Query
            Condition = Query.Condition
            Or = Query.Or

            # Create a query object.
            columns = "d.id", "d.title"
            query = Query("document d", *columns).order("d.title")

            # Add the conditions: one for each field with a value.
            conditions = []
            have_doctype = False
            for field in self.fields:

                # See if we got a value for this field.
                value = field.var.strip() if field.var else None
                if not value:
                    continue

                # Remember the criterion for display to the user.
                self.criteria.append(value)

                # If 'selectors' is a string, it's a column in `document`.
                value_op = self.__getQueryOp(value)
                if isinstance(field.selectors, str):
                    column = f"d.{field.selectors}"
                    conditions.append(Condition(column, value, value_op))
                    continue

                # Build up a subquery for the field.
                have_doctype = True
                subquery = Query("query_term", "doc_id").unique()

                # These are ORd together.
                selector_conditions = []
                for path in field.selectors:
                    path_op = "LIKE" if "%" in path else "="

                    # Simple case: test for a string stored in the doc.
                    if not path.endswith("/@cdr:ref[int_val]"):
                        path_test = Condition("path", path, path_op)
                        value_test = Condition("value", value, value_op)
                        selector_conditions.append((path_test, value_test))

                    # Trickier case: find values in linked documents.
                    else:
                        path = path.replace("[int_val]", "")
                        title_query = Query("document", "id").unique()
                        args = "title", value, value_op
                        title_query.where(Condition(*args))
                        args = "int_val", title_query, "IN"
                        title_test = Condition(*args)
                        path_test = Condition("path", path, path_op)
                        selector_conditions.append((path_test, title_test))

                # Add the conditions for this field's selectors to the mix.
                subquery.where(Or(*selector_conditions))
                conditions.append(Condition("d.id", subquery, "IN"))

            # Sanity check.
            if not conditions:
                raise Exception("No search conditions specified")

            # If all the selectors are from the document table (or,
            # view, to be precise), then we still need to narrow
            # the search to this document type.
            if not have_doctype:
                query.join("doc_type t", "t.id = d.doc_type")
                query.where(Condition("t.name", self.doctype))

            # Plug the top-level conditions into the query.
            if self.match_all:
                for condition in conditions:
                    query.where(condition)
            else:
                query.where(Or(*conditions))

            # All the fields with values have been folded into the query.
            return query

        @staticmethod
        def __getQueryOp(query):
            """Determine whether query contains unescaped wildcards.

            Required positional argument:
            query - string contining search query

            Return:
            SQL operator "LIKE" or "="
            """
            escaped = 0
            for char in query:
                if char == '\\':
                    escaped = not escaped
                elif not escaped and char in "_%":
                    return "LIKE"
            return "="
